import openpyxl as op
import sys
from PyQt5 import QtGui, QtWidgets
import Unprofitability
from mainDesign import Ui_MainWindow
from pathlib import Path
import Sales_report
import Form_0420162
import Court
import Triangles

class App(QtWidgets.QMainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        self.setFixedSize(self.size())
        self.setWindowIcon(QtGui.QIcon('icon.png'))

        self.ui.salefoldButton.clicked.connect(self.getfolder_sales)
        self.ui.unafoldButton.clicked.connect(self.getfolder_una)
        self.ui.formfoldButton.clicked.connect(self.getfolder_162)
        self.ui.courtfoldButton.clicked.connect(self.getfolder_court)
        self.ui.trefoldButton.clicked.connect(self.getfolder_triangles)

        self.ui.saleButton.clicked.connect(self.start_sales)
        self.ui.unaButton.clicked.connect(self.start_una)
        self.ui.form162Button.clicked.connect(self.start_162)
        self.ui.courtButton.clicked.connect(self.start_court)
        self.ui.treButton.clicked.connect(self.start_triangles)

        self.ui.saleSaveButton.clicked.connect(self.save_sale)
        self.ui.unaSaveButton.clicked.connect(self.save_una)
        self.ui.formSaveButton.clicked.connect(self.save_162)
        self.ui.courtSaveButton.clicked.connect(self.save_court)
        self.ui.treSaveButton.clicked.connect(self.save_triangles)



    def err_mes(self, err):
        error = QtWidgets.QMessageBox()
        error.setWindowTitle("Ошибка")
        error.setText(err)
        error.setIcon(QtWidgets.QMessageBox.Warning)
        error.setStandardButtons(QtWidgets.QMessageBox.Ok)
        error.exec()


    def start_sales(self):
        self.ui.saleButton.setEnabled(0)
        if (Path(self.ui.lineEdit.text()).is_dir()):
                sales = Sales_report.Sales_report(self.ui.lineEdit.text())
                try:
                    sales.form()
                except Exception:
                    self.err_mes(sales.error)
        else:
                self.err_mes('Укажите корректный путь к папке с файлами')
        self.ui.saleButton.setEnabled(1)

    def start_una(self):
        self.ui.unaButton.setEnabled(0)
        if (Path(self.ui.lineEdit_2.text()).is_dir()):
                una = Unprofitability.Unprofitability(self.ui.lineEdit_2.text())
                try:
                    una.form()
                except Exception:
                    self.err_mes(una.error)
        else:
                self.err_mes('Укажите корректный путь к папке с файлами')
        self.ui.unaButton.setEnabled(1)

    def start_162(self):
        self.ui.form162Button.setEnabled(0)
        if (Path(self.ui.lineEdit_3.text()).is_dir()):
                form162 = Form_0420162.Form_0420162(self.ui.lineEdit_3.text())
                try:
                    form162.form()
                except Exception:
                    self.err_mes(form162.error)
        else:
                self.err_mes('Укажите корректный путь к папке с файлами')
        self.ui.form162Button.setEnabled(1)

    def start_court(self):
        self.ui.courtButton.setEnabled(0)

        if (Path(self.ui.lineEdit_4.text()).is_dir()):
                court = Court.Court(self.ui.lineEdit_4.text())
                try:
                    court.form()
                except Exception:
                    self.err_mes(court.error)
        else:
            self.err_mes('Укажите корректный путь к папке с файлами')
        self.ui.courtButton.setEnabled(1)

    def start_triangles(self):
        self.ui.treButton.setEnabled(0)
        if (Path(self.ui.lineEdit_5.text()).is_dir()):
                triangles = Triangles.Triangles(self.ui.lineEdit_5.text())
                try:
                    triangles.form()
                except Exception:
                    self.err_mes(triangles.error)
        else:
            self.err_mes('Укажите корректный путь к папке с файлами')
        self.ui.treButton.setEnabled(1)

    def getfolder_sales(self):
        self.folder = QtWidgets.QFileDialog.getExistingDirectory(self, "Выбрать папку с файлами")
        self.ui.lineEdit.setText(self.folder)

    def getfolder_una(self):
        self.folder = QtWidgets.QFileDialog.getExistingDirectory(self, "Выбрать папку с файлами")
        self.ui.lineEdit_2.setText(self.folder)

    def getfolder_162(self):
        self.folder = QtWidgets.QFileDialog.getExistingDirectory(self, "Выбрать папку с файлами")
        self.ui.lineEdit_3.setText(self.folder)

    def getfolder_court(self):
        self.folder = QtWidgets.QFileDialog.getExistingDirectory(self, "Выбрать папку с файлами")
        self.ui.lineEdit_4.setText(self.folder)

    def getfolder_triangles(self):
        self.folder = QtWidgets.QFileDialog.getExistingDirectory(self, "Выбрать папку с файлами")
        self.ui.lineEdit_5.setText(self.folder)

    def save_sale(self):
        try:
            self.file = QtWidgets.QFileDialog.getSaveFileName(self, "Сохранить файл",
                                                              '', '.xlsx')
            wb = op.load_workbook('!Отчет по продажам.xlsx')
            wb.save(self.file[0]+'.xlsx')
        except Exception:
            self.err_mes('Отчет не сформирован')

    def save_una(self):
        try:
            self.file = QtWidgets.QFileDialog.getSaveFileName(self, "Сохранить файл",
                                                              '', '.xlsx')
            wb = op.load_workbook('!Убыточность.xlsx')
            wb.save(self.file[0]+'.xlsx')
        except Exception:
            self.err_mes('Отчет не сформирован')

    def save_162(self):
        try:
            self.file = QtWidgets.QFileDialog.getSaveFileName(self, "Сохранить файл",
                                                              '', '.xlsx')
            wb = op.load_workbook('!Форма 162.xlsx')
            wb.save(self.file[0]+'.xlsx')
        except Exception:
            self.err_mes('Отчет не сформирован')

    def save_court(self):
        try:
            self.file = QtWidgets.QFileDialog.getSaveFileName(self, "Сохранить файл",
                                                              '', '.xlsx')
            wb = op.load_workbook('!Судебная Отчетность.xlsx')
            wb.save(self.file[0]+'.xlsx')
        except Exception:
            self.err_mes('Отчет не сформирован')

    def save_triangles(self):
        try:
            self.file = QtWidgets.QFileDialog.getSaveFileName(self, "Сохранить файл",
                                                              '', '.xlsx')
            wb = op.load_workbook('!Треугольники.xlsx')
            wb.save(self.file[0]+'.xlsx')
        except Exception:
            self.err_mes('Отчет не сформирован')

if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    window = App()
    window.show()
    app.exit(app.exec())