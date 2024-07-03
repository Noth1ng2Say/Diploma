import os
import pandas as pd
import openpyxl as op

class Court:
    error = ''
    def __init__(self, folder):
        self.folder = folder



    def form(self):

        directory = self.folder
        files = os.listdir(directory)
        required_files = ['ВыплатыТрадиционные.xlsx', 'ВыплатаПотерпевшимПВУ.xlsx', 'Накладные расходы 71410.xlsx', 'Накладные расходы 71418.xlsx']

        if (set(required_files) != set(required_files) & set(files)):
            differences = list(set(required_files) - set(files))
            append_or_crush = ", ".join(str(element) for element in differences)
            self.error = f'В директории не хватает файлов: {append_or_crush}'
            raise Exception

        payment=self.folder + '/ВыплатыТрадиционные.xlsx'
        pay_pvu=self.folder + '/ВыплатаПотерпевшимПВУ.xlsx'
        overheads=self.folder + '/Накладные расходы 71410.xlsx'
        overheads_pvu=self.folder + '/Накладные расходы 71418.xlsx'
        df_pay = pd.read_excel(payment)
        df_pvu = pd.read_excel(pay_pvu, skiprows=1)
        df_overh = pd.read_excel(overheads, skiprows=1)
        df_over_pvu = pd.read_excel(overheads_pvu, skiprows=1)
        df_pay.drop([0, 1, 2], inplace=True)

        columns = df_pay.columns.tolist()
        required_col = ['Признак ПВУ', 'Вид страхования', 'ОКАТО территории использо- вания ТС Первоначальный',
                        'По реше- нию суда', "Суммы выплаты", 'Страховой случай']

        if (set(required_col) != set(required_col) & set(columns)):
            differences = list(set(required_col) - set(columns))
            append_or_crush = ", ".join(str(element) for element in differences)
            self.error = f'В таблице ВыплатыТрадиционные не хватает столбцов:\n{append_or_crush}'
            raise Exception

        columns = df_pvu.columns.tolist()
        required_col = ['ТерриторияИспользованияТСКодОКАТО', 'Сумма', 'Заявка', 'ДоплатаПоРешениюСуда']

        if (set(required_col) != set(required_col) & set(columns)):
            differences = list(set(required_col) - set(columns))
            append_or_crush = ", ".join(str(element) for element in differences)
            self.error = f'В таблице ВыплатаПотерпевшимПВУ не хватает столбцов:\n{append_or_crush}'
            raise Exception

        columns = df_over_pvu.columns.tolist()
        required_col = ['СчетДт', 'СубконтоДт2', 'СубконтоДт1', 'КодОКАТО_ТерриторияИспользованияТС', 'Сумма']

        if (set(required_col) != set(required_col) & set(columns)):
            differences = list(set(required_col) - set(columns))
            append_or_crush = ", ".join(str(element) for element in differences)
            self.error = f'В таблице Накладные расходы 71418 не хватает столбцов:\n{append_or_crush}'
            raise Exception

        columns = df_overh.columns.tolist()

        if (set(required_col) != set(required_col) & set(columns)):
            differences = list(set(required_col) - set(columns))
            append_or_crush = ", ".join(str(element) for element in differences)
            self.error = f'В таблице Накладные расходы 71410 не хватает столбцов:\n{append_or_crush}'
            raise Exception

        df_pay = df_pay[df_pay['Признак ПВУ'] != 'Да']
        df_pay = df_pay[df_pay['Вид страхования'] == 'Обязательное страхование гражданской ответственности владельцев транспортных средств']
        df_pay.insert(61, 'ОКАТО', df_pay['ОКАТО территории использо- вания ТС Первоначальный']/1000000000)
        df_pay['ОКАТО'] = df_pay['ОКАТО'].astype('int')

        df_pvu = df_pvu.dropna(subset=['Сумма'])
        df_pvu.insert(4, 'ОКАТО', df_pvu['ТерриторияИспользованияТСКодОКАТО']/1000000000)
        df_pvu['ОКАТО'] = df_pvu['ОКАТО'].astype('int')

        df_overh = df_overh.dropna(subset=['Сумма'])
        df_overh=df_overh[df_overh['СчетДт'] == 71410]
        df_overh=df_overh[df_overh['СубконтоДт2'] == 'Обязательное страхование гражданской ответственности владельцев транспортных средств']
        df_overh = df_overh[df_overh['СубконтоДт1'].str.match(r'.*(исполнительные листы)')]
        df_overh.insert(4, 'ОКАТО', df_overh['КодОКАТО_ТерриторияИспользованияТС']/1000000000)
        df_overh['ОКАТО'] = df_overh['ОКАТО'].astype('int')

        msn = ['Расходы на оплату компенсации морального вреда (исполнительные листы)', 'Расходы  неустоек (исполнительные листы)', 'Расходы на оплату компенсации морального вреда ПВУ (исполнительные листы)']
        def type_of_overheads(row):
            if (row['СубконтоДт1'] in msn):
                value = 'МШН'
            else:
                value =  'Прочее'
            return value

        df_overh.insert(5, 'мшн', df_overh.apply(type_of_overheads, axis=1))

        df_over_pvu = df_over_pvu.dropna(subset=['Сумма'])
        df_over_pvu=df_over_pvu[df_over_pvu['СчетДт'] == 71418]
        df_over_pvu = df_over_pvu[df_over_pvu['СубконтоДт1'].str.match(r'.*(исполнительные листы)')]
        df_over_pvu.insert(4, 'ОКАТО', df_over_pvu['КодОКАТО_ТерриторияИспользованияТС']/1000000000)
        df_over_pvu['ОКАТО'] = df_over_pvu['ОКАТО'].astype('int')
        df_over_pvu.insert(5, 'мшн', df_over_pvu.apply(type_of_overheads, axis=1))


        table_pay = pd.pivot_table(df_pay,
                       index=["ОКАТО"],
                       columns = ['По реше- нию суда'],
                       values=["Суммы выплаты", 'Страховой случай'],
                       aggfunc={"Суммы выплаты": pd.Series.sum, "Страховой случай": pd.Series.nunique},
                        fill_value=0).reset_index()


        table_pvu = pd.pivot_table(df_pvu,
                       index=["ОКАТО"],
                       columns = ['ДоплатаПоРешениюСуда'],
                       values=["Сумма", "Заявка"],
                       aggfunc= {"Сумма": pd.Series.sum, "Заявка": pd.Series.nunique},
                        fill_value=0).reset_index()


        table_overheads = pd.pivot_table(df_overh,
                       index=["ОКАТО"],
                       columns = ['мшн'],
                       values=["Сумма"],
                       aggfunc=pd.Series.sum,
                        fill_value=0,
                        margins=True).reset_index()


        table_overheads_pvu = pd.pivot_table(df_over_pvu,
                       index=["ОКАТО"],
                       columns = ['мшн'],
                       values=["Сумма"],
                       aggfunc=pd.Series.sum,
                        fill_value=0,
                        margins=True).reset_index()


        report = self.folder + '/Форма_Выплаты по судебным решениям.xlsx'
        wb = op.load_workbook(report)
        sheet = wb.active


        sheet['F5'] = 'АО СК "Армеец"'
        for index, row in table_pay.iterrows():
            for i in range (13, 100):
                code = int(sheet.cell(row = i, column = 2).value)
                if(code == row.values[0]):
                    sheet.cell(row=i, column=3).value = row.values[2]
                    sheet.cell(row=i, column=4).value = row.values[4]
                    sheet.cell(row=i, column=8).value = row.values[1]
                    sheet.cell(row=i, column=9).value = row.values[3]


        for index, row in table_pvu.iterrows():
            for i in range (13, 100):
                code = int(sheet.cell(row = i, column = 2).value)
                if(code == row.values[0]):
                    sheet.cell(row=i, column=13).value = row.values[2]
                    sheet.cell(row=i, column=14).value = row.values[4]
                    sheet.cell(row=i, column=18).value = row.values[1]
                    sheet.cell(row=i, column=19).value = row.values[3]


        for index, row in table_overheads.iterrows():
            for i in range (13, 100):
                code = int(sheet.cell(row = i, column = 2).value)
                if(code == row.values[0]):
                    sheet.cell(row=i, column=10).value = round(row.values[3])
                    sheet.cell(row=i, column=11).value = row.values[1]
                    sheet.cell(row=i, column=12).value = row.values[2]

        for index, row in table_overheads_pvu.iterrows():
            for i in range (13, 100):
                code = int(sheet.cell(row = i, column = 2).value)
                if(code == row.values[0]):
                    sheet.cell(row=i, column=20).value = round(row.values[3])
                    sheet.cell(row=i, column=21).value = row.values[1]
                    sheet.cell(row=i, column=22).value = row.values[2]


        pre_trial = self.folder + '/УрегулированиеФУ.xlsx'
        nwb = op.load_workbook(pre_trial)
        nsheet = nwb.active

        for i in range(9, 96):
            for j in range (13, 100):
                if (int(sheet.cell(row = j, column = 2).value) == int(nsheet.cell(row = i, column = 2).value)):
                    sheet.cell(row=j, column=5).value = nsheet.cell(row = i, column = 3).value
                    sheet.cell(row=j, column=6).value = nsheet.cell(row=i, column=4).value
                    sheet.cell(row=j, column=7).value = nsheet.cell(row=i, column=5).value
                    sheet.cell(row=j, column=15).value = nsheet.cell(row=i, column=6).value
                    sheet.cell(row=j, column=16).value = nsheet.cell(row=i, column=7).value
                    sheet.cell(row=j, column=17).value = nsheet.cell(row=i, column=8).value


        wb.save('!Судебная Отчетность.xlsx')
        import subprocess
        subprocess.call('!Судебная Отчетность.xlsx', shell=True)

