import os
import pandas as pd
import openpyxl as op
import re
from dateutil.relativedelta import relativedelta

class Triangles:
    def __init__(self, folder):
        self.folder = folder


    def form(self):
        #Выплаты
        directory = self.folder
        files = os.listdir(directory)
        required_files = ['Выплаты 2 плавающих года.xlsx', 'Заявления 2 плавающих.xlsx']

        if (sorted(list(set(required_files))) != sorted(list(set(required_files) & set(files)))):
            differences = list(set(required_files) - set(files))
            append_or_crush = ", ".join(str(element) for element in differences)
            self.error = f'В директории не хватает файлов: {append_or_crush}'
            raise Exception

        payment=self.folder + '/Выплаты 2 плавающих года.xlsx'
        df_pay = pd.read_excel(payment)
        df_pay.drop([0, 1, 2], inplace=True)

        columns = df_pay.columns.tolist()
        required_col = ['Период', 'Дата начисления полиса', 'Подразделение первоначального полиса', 'Номер первоначального полиса', 'ОКАТО территории использования ТС первоначального полиса', 'Суммы выплаты', 'Страховой случай']

        if (sorted(list(set(required_col))) != sorted(list(set(required_col) & set(columns)))):
            differences = list(set(required_col) - set(columns))
            append_or_crush = ", ".join(str(element) for element in differences)
            self.error = f'В таблице Выплаты 2 плавающих года не хватает столбцов:\n{append_or_crush}'
            raise Exception

        statements = self.folder + '/Заявления 2 плавающих.xlsx'
        df_stat = pd.read_excel(statements)
        df_stat.drop(0, inplace=True)

        columns = df_stat.columns.tolist()
        required_col = ['Дата убытка', 'Дата первоначального полиса', 'Подразделение первоначального полиса',
                        'Номер первоначального полиса', 'ОКАТО территории использования ТС первоначального полиса',
                        'Страховой случай']

        if (sorted(list(set(required_col))) != sorted(list(set(required_col) & set(columns)))):
            differences = list(set(required_col) - set(columns))
            append_or_crush = ", ".join(str(element) for element in differences)
            self.error = f'В таблице Заявления 2 плавающих не хватает столбцов:\n{append_or_crush}'
            raise Exception


        date_cols = ['Период', 'Дата начисления полиса']
        df_pay[date_cols] = df_pay[date_cols].apply(lambda x: pd.to_datetime(x, format="%d.%m.%Y"))
        max_date = df_pay['Период'].max() + relativedelta(day=31)
        min_date = df_pay['Период'].min() - relativedelta(day=1)
        df_pay.insert(2, 'Год и месяц оплаты', df_pay['Период'].dt.to_period("M"))
        df_pay.insert(41, 'Год и месяц полиса', df_pay['Дата начисления полиса'].apply(lambda x: x.to_period("M") if (x > max_date - relativedelta(years=2)) else '0'))
        df_pay = df_pay[df_pay['Год и месяц полиса'] != '0']

        period = df_pay['Год и месяц оплаты'].unique()
        period_for_xl = []
        for i in period:
           period_for_xl.append(i.to_timestamp())

        def Sales_channel(row):
            if (row['Подразделение первоначального полиса'] == 'Точка продаж ЕГАРАНТ' or row['Подразделение первоначального полиса'] == 'Агент РСА (ЦО)'):
                value = 'Е-Гарант'
            elif (re.fullmatch(r'ХХХ.*', row['Номер первоначального полиса'])):
               value =  'Интернет'
            elif(row['ОКАТО территории использования ТС первоначального полиса']/1000000000 == 3 ):
                value =  'Краснодар'
            elif (row['ОКАТО территории использования ТС первоначального полиса'] / 1000000000 == 92):
                value = 'Татарстан'
            else:
                value = 'Прочие'
            return value

        df_pay.insert(3, 'Канал продаж', df_pay.apply(Sales_channel, axis=1))


        table_pay = pd.pivot_table(df_pay,
                       index= ['Год и месяц полиса'],
                       columns = ['Год и месяц оплаты'],
                       values=['Суммы выплаты'] ,
                       aggfunc= pd.Series.sum,
                        fill_value=0).reset_index()
        table_pay_internet= pd.pivot_table(df_pay[df_pay['Канал продаж'] == 'Интернет'],
                       index= ['Год и месяц полиса'],
                       columns = ['Год и месяц оплаты'],
                       values=['Суммы выплаты'] ,
                       aggfunc= pd.Series.sum,
                        fill_value=0).reset_index()
        table_pay_garant= pd.pivot_table(df_pay[df_pay['Канал продаж'] == 'Е-Гарант'],
                       index= ['Год и месяц полиса'],
                       columns = ['Год и месяц оплаты'],
                       values=['Суммы выплаты'] ,
                       aggfunc= pd.Series.sum,
                        fill_value=0).reset_index()
        table_pay_03= pd.pivot_table(df_pay[df_pay['Канал продаж'] == 'Краснодар'],
                       index= ['Год и месяц полиса'],
                       columns = ['Год и месяц оплаты'],
                       values=['Суммы выплаты'] ,
                       aggfunc= pd.Series.sum,
                        fill_value=0).reset_index()
        table_pay_92= pd.pivot_table(df_pay[df_pay['Канал продаж'] == 'Татарстан'],
                       index= ['Год и месяц полиса'],
                       columns = ['Год и месяц оплаты'],
                       values=['Суммы выплаты'] ,
                       aggfunc= pd.Series.sum,
                        fill_value=0).reset_index()
        table_pay_other= pd.pivot_table(df_pay[df_pay['Канал продаж'] == 'Прочие'],
                       index= ['Год и месяц полиса'],
                       columns = ['Год и месяц оплаты'],
                       values=['Суммы выплаты'] ,
                       aggfunc= pd.Series.sum,
                        fill_value=0).reset_index()
        table_pay_count = pd.pivot_table(df_pay,
                       index= ['Год и месяц полиса'],
                       columns = ['Год и месяц оплаты'],
                       values=['Страховой случай'] ,
                       aggfunc= pd.Series.nunique,
                        fill_value=0).reset_index()
        table_pay_count_internet = pd.pivot_table(df_pay[df_pay['Канал продаж'] == 'Интернет'],
                       index= ['Год и месяц полиса'],
                       columns = ['Год и месяц оплаты'],
                       values=['Страховой случай'] ,
                       aggfunc= pd.Series.nunique,
                        fill_value=0).reset_index()
        table_pay_count_garant = pd.pivot_table(df_pay[df_pay['Канал продаж'] == 'Е-Гарант'],
                       index= ['Год и месяц полиса'],
                       columns = ['Год и месяц оплаты'],
                       values=['Страховой случай'] ,
                       aggfunc= pd.Series.nunique,
                        fill_value=0).reset_index()
        table_pay_count_03 = pd.pivot_table(df_pay[df_pay['Канал продаж'] == 'Краснодар'],
                       index= ['Год и месяц полиса'],
                       columns = ['Год и месяц оплаты'],
                       values=['Страховой случай'] ,
                       aggfunc= pd.Series.nunique,
                        fill_value=0).reset_index()
        table_pay_count_92 = pd.pivot_table(df_pay[df_pay['Канал продаж'] == 'Татарстан'],
                       index= ['Год и месяц полиса'],
                       columns = ['Год и месяц оплаты'],
                       values=['Страховой случай'] ,
                       aggfunc= pd.Series.nunique,
                        fill_value=0).reset_index()
        table_pay_count_other = pd.pivot_table(df_pay[df_pay['Канал продаж'] == 'Прочие'],
                       index= ['Год и месяц полиса'],
                       columns = ['Год и месяц оплаты'],
                       values=['Страховой случай'] ,
                       aggfunc= pd.Series.nunique,
                        fill_value=0).reset_index()

        #заяленные
        date_cols = ['Дата убытка', 'Дата первоначального полиса']
        df_stat[date_cols] = df_stat[date_cols].apply(lambda x: pd.to_datetime(x, format="%d.%m.%Y"))
        max_date = df_stat['Дата убытка'].max() + relativedelta(day=31)
        df_stat.insert(2, 'Год и месяц заявления', df_stat['Дата убытка'].dt.to_period("M"))
        df_stat.insert(41, 'Год и месяц полиса', df_stat['Дата первоначального полиса'].apply(lambda x: x.to_period("M") if (x > max_date - relativedelta(years=2)) else '0'))
        df_stat = df_stat[df_stat['Год и месяц полиса'] != '0']

        df_stat.insert(4, 'Канал продаж', df_stat.apply(Sales_channel, axis=1))

        table_stat = pd.pivot_table(df_stat,
                       index= ['Год и месяц полиса'],
                       columns = ['Год и месяц заявления'],
                       values=['Страховой случай'] ,
                       aggfunc= pd.Series.nunique,
                        fill_value=0).reset_index()
        table_stat_internet= pd.pivot_table(df_stat[df_stat['Канал продаж'] == 'Интернет'],
                       index= ['Год и месяц полиса'],
                       columns = ['Год и месяц заявления'],
                       values=['Страховой случай'] ,
                       aggfunc= pd.Series.nunique,
                        fill_value=0).reset_index()
        table_stat_garant= pd.pivot_table(df_stat[df_stat['Канал продаж'] == 'Е-Гарант'],
                       index= ['Год и месяц полиса'],
                       columns = ['Год и месяц заявления'],
                       values=['Страховой случай'] ,
                       aggfunc= pd.Series.nunique,
                        fill_value=0).reset_index()
        table_stat_03= pd.pivot_table(df_stat[df_stat['Канал продаж'] == 'Краснодар'],
                       index= ['Год и месяц полиса'],
                       columns = ['Год и месяц заявления'],
                       values=['Страховой случай'] ,
                       aggfunc= pd.Series.nunique,
                        fill_value=0).reset_index()
        table_stat_92= pd.pivot_table(df_stat[df_stat['Канал продаж'] == 'Татарстан'],
                       index= ['Год и месяц полиса'],
                       columns = ['Год и месяц заявления'],
                       values=['Страховой случай'] ,
                       aggfunc= pd.Series.nunique,
                        fill_value=0).reset_index()
        table_stat_other= pd.pivot_table(df_stat[df_stat['Канал продаж'] == 'Прочие'],
                       index= ['Год и месяц полиса'],
                       columns = ['Год и месяц заявления'],
                       values=['Страховой случай'] ,
                       aggfunc= pd.Series.nunique,
                        fill_value=0).reset_index()

        table_stat.to_excel('Проверка.xlsx')

        #Заполнение
        form = 'Форма Треугольники.xlsx'
        wb = op.load_workbook(form)
        all = wb['Всего']
        garant = wb['Единый агент_Е-гарант']
        internet = wb['Интернет']
        krasnodar = wb['Краснодарский край']
        tatar = wb['Республика Татарстан']
        other = wb['Прочие продажи']
        sheets = [all, garant, internet, krasnodar, tatar, other]

        num = [0, 30, 58]
        for i in sheets:
         i['B5'] = f'Сведения за период с {min_date.strftime("%d.%m.%Y")} по {max_date.strftime("%d.%m.%Y")} о суммах и количестве заявленных и оплаченных Страховщиком убытков по договорам обязательного страхования гражданской ответственности владельцев транспортных средств (ОСАГО)'
         for n in num:
             for j in range(9, 33):
                i.cell(row=j+n, column=1).value = period_for_xl[j-9]
             for j in range(2, 26):
                i.cell(row=8+n, column=j).value = period_for_xl[j - 2]


        #Всего
        table_pay.columns = table_pay.columns.droplevel(0)
        table_pay_count.columns = table_pay_count.columns.droplevel(0)
        col = table_pay.columns.tolist()
        col.pop(0)

        for index, row in table_pay.iterrows():
            for i in range (9, 33):
                for n in range(0, len(col)):
                    for j in range(2, 26):
                        if((all.cell(row = i, column = 1).value).to_period("M") == row.values[0] and (all.cell(row = 8, column = j).value).to_period("M") == col[n]):
                            all.cell(row = i, column = j).value = row.values[n+1] / 1000
                            break

        for index, row in table_pay_count.iterrows():
            for i in range (39, 63):
                for n in range(0, len(col)):
                    for j in range(2, 26):
                        if((all.cell(row = i, column = 1).value).to_period("M") == row.values[0] and (all.cell(row = 8, column = j).value).to_period("M") == col[n]):
                            all.cell(row = i, column = j).value = row.values[n+1]
                            break

        table_stat.columns = table_stat.columns.droplevel(0)
        col = table_stat.columns.tolist()
        col.pop(0)

        for index, row in table_stat.iterrows():
            for i in range (67, 91):
                for n in range(0, len(col)):
                    for j in range(2, 26):
                        if((all.cell(row = i, column = 1).value).to_period("M") == row.values[0] and (all.cell(row = 8, column = j).value).to_period("M") == col[n]):
                            all.cell(row = i, column = j).value = row.values[n+1]
                            break

        #Е-Гарант
        table_pay_garant.columns = table_pay_garant.columns.droplevel(0)
        table_pay_count_garant.columns = table_pay_count_garant.columns.droplevel(0)
        col = table_pay_garant.columns.tolist()
        col.pop(0)

        for index, row in table_pay_garant.iterrows():
            for i in range (9, 33):
                for n in range(0, len(col)):
                    for j in range(2, 26):
                        if((garant.cell(row = i, column = 1).value).to_period("M") == row.values[0] and (garant.cell(row = 8, column = j).value).to_period("M") == col[n]):
                            garant.cell(row = i, column = j).value = row.values[n+1] / 1000
                            break

        for index, row in table_pay_count_garant.iterrows():
            for i in range (39, 63):
                for n in range(0, len(col)):
                    for j in range(2, 26):
                        if((garant.cell(row = i, column = 1).value).to_period("M") == row.values[0] and (garant.cell(row = 8, column = j).value).to_period("M") == col[n]):
                            garant.cell(row = i, column = j).value = row.values[n+1]
                            break

        table_stat_garant.columns = table_stat_garant.columns.droplevel(0)
        col = table_stat_garant.columns.tolist()
        col.pop(0)

        for index, row in table_stat_garant.iterrows():
            for i in range (67, 91):
                for n in range(0, len(col)):
                    for j in range(2, 26):
                        if ((garant.cell(row=i, column=1).value).to_period("M") == row.values[0] and (garant.cell(row=8, column=j).value).to_period("M") == col[n]):
                            garant.cell(row=i, column=j).value = row.values[n + 1]
                            break

        #Интернет
        table_pay_internet.columns = table_pay_internet.columns.droplevel(0)
        table_pay_count_internet.columns = table_pay_count_internet.columns.droplevel(0)
        col = table_pay_internet.columns.tolist()
        col.pop(0)

        for index, row in table_pay_internet.iterrows():
            for i in range (9, 33):
                for n in range(0, len(col)):
                    for j in range(2, 26):
                        if((internet.cell(row = i, column = 1).value).to_period("M") == row.values[0] and (internet.cell(row = 8, column = j).value).to_period("M") == col[n]):
                            internet.cell(row = i, column = j).value = row.values[n+1] / 1000
                            break

        for index, row in table_pay_count_internet.iterrows():
            for i in range (39, 63):
                for n in range(0, len(col)):
                    for j in range(2, 26):
                        if((internet.cell(row = i, column = 1).value).to_period("M") == row.values[0] and (internet.cell(row = 8, column = j).value).to_period("M") == col[n]):
                            internet.cell(row = i, column = j).value = row.values[n+1]
                            break

        table_stat_internet.columns = table_stat_internet.columns.droplevel(0)
        col = table_stat_internet.columns.tolist()
        col.pop(0)

        for index, row in table_stat_internet.iterrows():
            for i in range (67, 91):
                for n in range(0, len(col)):
                    for j in range(2, 26):
                        if ((internet.cell(row=i, column=1).value).to_period("M") == row.values[0] and (internet.cell(row=8, column=j).value).to_period("M") == col[n]):
                            internet.cell(row=i, column=j).value = row.values[n + 1]
                            break

        #Краснодар
        table_pay_03.columns = table_pay_03.columns.droplevel(0)
        table_pay_count_03.columns = table_pay_count_03.columns.droplevel(0)
        col = table_pay_03.columns.tolist()
        col.pop(0)

        for index, row in table_pay_03.iterrows():
            for i in range (9, 33):
                for n in range(0, len(col)):
                    for j in range(2, 26):
                        if((krasnodar.cell(row = i, column = 1).value).to_period("M") == row.values[0] and (krasnodar.cell(row = 8, column = j).value).to_period("M") == col[n]):
                            krasnodar.cell(row = i, column = j).value = row.values[n+1] / 1000
                            break

        for index, row in table_pay_count_03.iterrows():
            for i in range (39, 63):
                for n in range(0, len(col)):
                    for j in range(2, 26):
                        if((krasnodar.cell(row = i, column = 1).value).to_period("M") == row.values[0] and (krasnodar.cell(row = 8, column = j).value).to_period("M") == col[n]):
                            krasnodar.cell(row = i, column = j).value = row.values[n+1]
                            break

        table_stat_03.columns = table_stat_03.columns.droplevel(0)
        col = table_stat_03.columns.tolist()
        col.pop(0)

        for index, row in table_stat_03.iterrows():
            for i in range (67, 91):
                for n in range(0, len(col)):
                    for j in range(2, 26):
                        if ((krasnodar.cell(row=i, column=1).value).to_period("M") == row.values[0] and (krasnodar.cell(row=8, column=j).value).to_period("M") == col[n]):
                            krasnodar.cell(row=i, column=j).value = row.values[n + 1]
                            break

        #Татарстан
        table_pay_92.columns = table_pay_92.columns.droplevel(0)
        table_pay_count_92.columns = table_pay_count_92.columns.droplevel(0)
        col = table_pay_92.columns.tolist()
        col.pop(0)

        for index, row in table_pay_92.iterrows():
            for i in range (9, 33):
                for n in range(0, len(col)):
                    for j in range(2, 26):
                        if((tatar.cell(row = i, column = 1).value).to_period("M") == row.values[0] and (tatar.cell(row = 8, column = j).value).to_period("M") == col[n]):
                            tatar.cell(row = i, column = j).value = row.values[n+1] / 1000
                            break

        for index, row in table_pay_count_92.iterrows():
            for i in range (39, 63):
                for n in range(0, len(col)):
                    for j in range(2, 26):
                        if((tatar.cell(row = i, column = 1).value).to_period("M") == row.values[0] and (tatar.cell(row = 8, column = j).value).to_period("M") == col[n]):
                            tatar.cell(row = i, column = j).value = row.values[n+1]
                            break

        table_stat_92.columns = table_stat_92.columns.droplevel(0)
        col = table_stat_92.columns.tolist()
        col.pop(0)

        for index, row in table_stat_92.iterrows():
            for i in range (67, 91):
                for n in range(0, len(col)):
                    for j in range(2, 26):
                        if ((tatar.cell(row=i, column=1).value).to_period("M") == row.values[0] and (tatar.cell(row=8, column=j).value).to_period("M") == col[n]):
                            tatar.cell(row=i, column=j).value = row.values[n + 1]
                            break

        #Прочие
        table_pay_other.columns = table_pay_other.columns.droplevel(0)
        table_pay_count_other.columns = table_pay_count_other.columns.droplevel(0)
        col = table_pay_other.columns.tolist()
        col.pop(0)

        for index, row in table_pay_other.iterrows():
            for i in range (9, 33):
                for n in range(0, len(col)):
                    for j in range(2, 26):
                        if((other.cell(row = i, column = 1).value).to_period("M") == row.values[0] and (other.cell(row = 8, column = j).value).to_period("M") == col[n]):
                            other.cell(row = i, column = j).value = row.values[n+1] / 1000
                            break

        for index, row in table_pay_count_other.iterrows():
            for i in range (39, 63):
                for n in range(0, len(col)):
                    for j in range(2, 26):
                        if((other.cell(row = i, column = 1).value).to_period("M") == row.values[0] and (other.cell(row = 8, column = j).value).to_period("M") == col[n]):
                            other.cell(row = i, column = j).value = row.values[n+1]
                            break

        table_stat_other.columns = table_stat_other.columns.droplevel(0)
        col = table_stat_other.columns.tolist()
        col.pop(0)

        for index, row in table_stat_other.iterrows():
            for i in range (67, 91):
                for n in range(0, len(col)):
                    for j in range(2, 26):
                        if ((other.cell(row=i, column=1).value).to_period("M") == row.values[0] and (other.cell(row=8, column=j).value).to_period("M") == col[n]):
                            other.cell(row=i, column=j).value = row.values[n + 1]
                            break

        wb.save('!Треугольники.xlsx')
        import subprocess
        subprocess.call('!Треугольники.xlsx', shell=True)

