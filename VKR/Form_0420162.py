import os
import numpy as np
import pandas as pd
import openpyxl as op
import re

class Form_0420162:
  error = ''
  def __init__(self, folder):
        self.folder = folder

  def form(self):
    directory = self.folder
    files = os.listdir(directory)
    required_files = ['Премии.xlsx', 'Заключенные.xlsx',  'Заявленные.xlsx', 'Отказы.xlsx', 'Убытки.xlsx', 'Комиссионное вознаграждение.xlsx', 'Перестрахование премии.xlsx', 'Перестрахование убытки.xlsx', 'Неустойки.xlsx',  'ЗНУ.xlsx']

    if (set(required_files) != set(required_files) & set(files)):
          differences = list(set(required_files) - set(files))
          append_or_crush = ", ".join(str(element) for element in differences)
          self.error = f'В директории не хватает файлов: {append_or_crush}'
          raise Exception

    insurance_type = self.folder + '/Вид страхования для статистики.xlsx'
    df_ins_type = pd.read_excel(insurance_type)

    #premium + досрочки
    premium = self.folder + '/Премии.xlsx'
    df_premium = pd.read_excel(premium, skiprows=2)
    df_premium = df_premium.drop(0)
    del df_premium['№']
    df_premium.insert(loc=len(df_premium.columns), column='Раздел', value=None)
    df_premium.insert(loc=len(df_premium.columns), column='Вид в 162', value=None)

    columns = df_premium.columns.tolist()
    required_col = ['Вид страхования', 'Код ОКАТО первоначального полиса', 'Подразделение первоначального полиса', 'Тип страхователя', 'Размер начисленной премии', 'Полис']

    if (set(required_col) != set(required_col) & set(columns)):
        differences = list(set(required_col) - set(columns))
        append_or_crush = ", ".join(str(element) for element in differences)
        self.error = f'В таблице Премии не хватает столбцов:\n{append_or_crush}'
        raise Exception

    df_ins_type = df_ins_type.set_index("ВидСтрахования")
    df_premium = df_premium.set_index("Вид страхования")
    df_premium.update(df_ins_type)
    df_premium = df_premium.reset_index()
    df_premium.insert(43, 'ОКАТО', np.where(df_premium[
                                            'Вид страхования'] != 'Обязательное страхование гражданской ответственности владельцев транспортных средств',
                                        df_premium['Код ОКАТО первоначального полиса'] / 1000,
                                        df_premium['Код ОКАТО первоначального полиса'] / 1000000000))
    df_premium['ОКАТО'] = df_premium['ОКАТО'].astype('int')


    def type_of_ins(row):
        if (row['Подразделение первоначального полиса'] == 'Точка продаж ЕГАРАНТ'):
            value = 'Е-Гарант'
        elif (re.fullmatch(r'ХХХ.*', row['Номер полиса'])):
            value = 'Электронный'
        else:
            value = 'Бумага'
        return value


    df_premium.insert(9, 'Вид договора', np.where(df_premium[
                                                  'Вид страхования'] == 'Обязательное страхование гражданской ответственности владельцев транспортных средств',
                                              df_premium.apply(type_of_ins, axis=1), None))

    df_premium_osago = df_premium.loc[df_premium[
                                      'Вид страхования'] == 'Обязательное страхование гражданской ответственности владельцев транспортных средств']
    df_termination = df_premium.loc[df_premium[
                                      'Статья начисления'] == 'Списание задолженности по премии дос. раст']

    table_section1_pre = pd.pivot_table(df_premium,
                                   index=["Раздел"],
                                   columns=['Тип страхователя'],
                                   values=["Размер начисленной премии"],
                                   aggfunc=pd.Series.sum,
                                   fill_value=0).reset_index()

    table_ins_type_pre = pd.pivot_table(df_premium,
                                    index=["Вид в 162"],
                                    columns=['Тип страхователя'],
                                    values=["Размер начисленной премии"],
                                    aggfunc=pd.Series.sum,
                                    fill_value=0).reset_index()

    table_okato_section_2_pre = pd.pivot_table(df_premium[df_premium[
                                                          'Вид страхования'] != 'Обязательное страхование гражданской ответственности владельцев транспортных средств'],
                                           index=["ОКАТО", "Раздел"],
                                           columns=['Тип страхователя'],
                                           values=["Размер начисленной премии"],
                                           aggfunc=pd.Series.sum,
                                           fill_value=0,
                                           margins=True).reset_index()

    table_section5_pre = pd.pivot_table(df_premium_osago,
                                   index=['ОКАТО', 'Тип страхователя'],
                                   columns=['Вид договора'],
                                   values=["Размер начисленной премии"],
                                   aggfunc=pd.Series.sum,
                                   fill_value=0).reset_index()

    table_section1_term = pd.pivot_table(df_termination,
                                   index=["Раздел"],
                                   columns=['Тип страхователя'],
                                   values=["Размер начисленной премии",  'Полис'],
                                   aggfunc={"Размер начисленной премии": pd.Series.sum,'Полис': pd.Series.nunique},
                                   fill_value=0).reset_index()

    table_section1_term_type = pd.pivot_table(df_termination,
                                   index=["Вид в 162"],
                                   columns=['Тип страхователя'],
                                   values=["Размер начисленной премии",  'Полис'],
                                   aggfunc={"Размер начисленной премии": pd.Series.sum,'Полис': pd.Series.nunique},
                                   fill_value=0).reset_index()

    table_section5_term = pd.pivot_table(df_termination[df_termination['Вид страхования'] == 'Обязательное страхование гражданской ответственности владельцев транспортных средств'],
                                   index=['ОКАТО', 'Тип страхователя'],
                                   columns=[],
                                   values={"Размер начисленной премии": pd.Series.sum,'Полис': pd.Series.nunique},
                                   aggfunc={"Размер начисленной премии": pd.Series.sum,'Полис': pd.Series.nunique},
                                   fill_value=0).reset_index()

    #Заключенные
    concluded_policies = self.folder + '/Заключенные.xlsx'
    df_concluded = pd.read_excel(concluded_policies, skiprows=2)
    df_concluded.drop(0, inplace=True)

    columns = df_concluded.columns.tolist()
    required_col = ['Вид страхования', 'Код ОКАТО первоначального полиса', 'Тип полиса', 'Технический доп',
                    'Страховая сумма', 'Тип страхователя', 'Размер начисленной премии', 'Количество застрахованных']

    if (set(required_col) != set(required_col) & set(columns)):
        differences = list(set(required_col) - set(columns))
        append_or_crush = ", ".join(str(element) for element in differences)
        self.error = f'В таблице Премии не хватает столбцов:\n{append_or_crush}'
        raise Exception

    df_concluded.insert(loc=len(df_concluded.columns), column='Раздел', value=None)
    df_concluded.insert(loc=len(df_concluded.columns), column='Вид в 162', value=None)
    df_concluded = df_concluded.set_index("Вид страхования")
    df_concluded.update(df_ins_type)
    df_concluded = df_concluded.reset_index()
    df_concluded.insert(43, 'ОКАТО2', np.where(df_concluded[
                                              'Вид страхования'] != 'Обязательное страхование гражданской ответственности владельцев транспортных средств',
                                          df_concluded['Код ОКАТО первоначального полиса'] / 1000,
                                          df_concluded['Код ОКАТО первоначального полиса'] / 1000000000))
    df_concluded['ОКАТО2'] = df_concluded['ОКАТО2'].astype('int')

    def policy_count(row):
        if (row['Размер начисленной премии'] < 0):
            value = 0
        elif (row['Тип полиса'] == 'Доп. соглашение'):
            value = 0
        elif (row['Технический доп'] == 'Да'):
            value = 0
        else:
            value = 1
        return value
    df_concluded.insert(3, 'Количество договоров', df_concluded.apply(policy_count, axis=1))

    df_concluded['Страховая сумма'] = df_concluded['Страховая сумма'].replace(r'\s+','',regex=True).replace(r'\,','.',regex=True).astype(float)
    def sum_insured(row):
        if (row['Вид страхования'] == 'Обязательное страхование гражданской ответственности владельцев транспортных средств' or row['Количество договоров'] == 0):
            value = 0
        else:
            value = row['Страховая сумма']
        return value
    df_concluded.insert(4, 'Страховая сумма2', df_concluded.apply(sum_insured, axis=1))

    df_concluded.insert(9, 'Вид договора', np.where(df_concluded[
                                                  'Вид страхования'] == 'Обязательное страхование гражданской ответственности владельцев транспортных средств',
                                              df_concluded.apply(type_of_ins, axis=1), None))

    table_section1_cluded = pd.pivot_table(df_concluded,
                                   index=["Раздел"],
                                   columns=['Тип страхователя'],
                                   values=["Размер начисленной премии", "Количество договоров", 'Страховая сумма2', 'Количество застрахованных'],
                                   aggfunc={"Размер начисленной премии": pd.Series.sum, "Количество договоров": pd.Series.sum, "Страховая сумма2": pd.Series.sum, "Количество застрахованных": pd.Series.sum},
                                   fill_value=0).reset_index()

    table_section1_cluded_type = pd.pivot_table(df_concluded,
                                   index=["Вид в 162"],
                                   columns=['Тип страхователя'],
                                   values=["Размер начисленной премии", "Количество договоров", 'Страховая сумма2', 'Количество застрахованных'],
                                   aggfunc={"Размер начисленной премии": pd.Series.sum, "Количество договоров": pd.Series.sum, "Страховая сумма2": pd.Series.sum, "Количество застрахованных": pd.Series.sum},
                                   fill_value=0).reset_index()

    table_section_2_cluded = pd.pivot_table(df_concluded[df_concluded['Вид страхования'] != 'Обязательное страхование гражданской ответственности владельцев транспортных средств'],
                                           index=["ОКАТО2", "Раздел"],
                                           columns=['Тип страхователя'],
                                           values=["Количество договоров", 'Страховая сумма2'],
                                           aggfunc={"Количество договоров": pd.Series.sum, "Страховая сумма2": pd.Series.sum},
                                           fill_value=0,
                                           margins=True).reset_index()

    table_section5_cluded = pd.pivot_table(df_concluded[df_concluded['Вид страхования'] == 'Обязательное страхование гражданской ответственности владельцев транспортных средств'],
                                   index=['ОКАТО2', 'Тип страхователя'],
                                   columns=['Вид договора'],
                                   values=["Размер начисленной премии", "Количество договоров"],
                                   aggfunc={"Количество договоров": pd.Series.sum, "Размер начисленной премии": pd.Series.sum},
                                   fill_value=0).reset_index()

    #Заявленные
    reported_accidents = self.folder+ '/Заявленные.xlsx'
    df_rep = pd.read_excel(reported_accidents)
    df_rep.drop(0, inplace=True)

    columns = df_rep.columns.tolist()
    required_col = ['Вид страхования', 'ОКАТО территории использования ТС первоначального полиса', 'Тип страхователя', 'Страховой случай']

    if (set(required_col) != set(required_col) & set(columns)):
        differences = list(set(required_col) - set(columns))
        append_or_crush = ", ".join(str(element) for element in differences)
        self.error = f'В таблице заявленные не хватает столбцов:\n{append_or_crush}'
        raise Exception

    df_rep.insert(loc=len(df_rep.columns), column='Раздел', value=None)
    df_rep.insert(loc=len(df_rep.columns), column='Вид в 162', value=None)
    df_rep = df_rep.set_index("Вид страхования")
    df_rep.update(df_ins_type)
    df_rep = df_rep.reset_index()
    df_rep.insert(3, 'ОКАТО', np.where(df_rep[
                                              'Вид страхования'] != 'Обязательное страхование гражданской ответственности владельцев транспортных средств',
                                          0,
                                          df_rep['ОКАТО территории использования ТС первоначального полиса'] / 1000000000))
    df_rep['ОКАТО'] = df_rep['ОКАТО'].astype('int')

    table_section1_rep = pd.pivot_table(df_rep,
                                   index=["Раздел"],
                                   columns=['Тип страхователя'],
                                   values=["Страховой случай"],
                                   aggfunc=pd.Series.nunique,
                                   fill_value=0).reset_index()

    table_section1_rep_type = pd.pivot_table(df_rep,
                                   index=["Вид в 162"],
                                   columns=['Тип страхователя'],
                                   values=["Страховой случай"],
                                   aggfunc=pd.Series.nunique,
                                   fill_value=0).reset_index()

    table_section5_rep = pd.pivot_table(df_rep[df_rep['Вид страхования'] == 'Обязательное страхование гражданской ответственности владельцев транспортных средств'],
                                   index=['ОКАТО', 'Тип страхователя'],
                                   columns=[],
                                   values=["Страховой случай"],
                                   aggfunc=pd.Series.nunique,
                                   fill_value=0).reset_index()

    #Отказы
    refusals = self.folder + '/Отказы.xlsx'
    df_refu = pd.read_excel(refusals)
    df_refu.drop(0, inplace=True)

    columns = df_refu.columns.tolist()
    required_col = ['Вид страхования', 'ОКАТО территории использования ТС первоначального полиса',
                    'Тип страхователя', 'Страховой случай']

    if (set(required_col) != set(required_col) & set(columns)):
        differences = list(set(required_col) - set(columns))
        append_or_crush = ", ".join(str(element) for element in differences)
        self.error = f'В таблице отказы не хватает столбцов:\n{append_or_crush}'
        raise Exception

    df_refu.insert(loc=len(df_refu.columns), column='Раздел', value=None)
    df_refu.insert(loc=len(df_refu.columns), column='Вид в 162', value=None)
    df_refu = df_refu.set_index("Вид страхования")
    df_refu.update(df_ins_type)
    df_refu = df_refu.reset_index()
    df_refu.insert(3, 'ОКАТО', np.where(df_refu[
                                              'Вид страхования'] != 'Обязательное страхование гражданской ответственности владельцев транспортных средств',
                                          0,
                                          df_refu['ОКАТО территории использования ТС первоначального полиса'] / 1000000000))
    df_refu['ОКАТО'] = df_refu['ОКАТО'].astype('int')

    table_section1_refu = pd.pivot_table(df_refu,
                                   index=["Раздел"],
                                   columns=['Тип страхователя'],
                                   values=["Страховой случай"],
                                   aggfunc=pd.Series.nunique,
                                   fill_value=0).reset_index()

    table_section1_refu_type = pd.pivot_table(df_refu,
                                   index=["Вид в 162"],
                                   columns=['Тип страхователя'],
                                   values=["Страховой случай"],
                                   aggfunc=pd.Series.nunique,
                                   fill_value=0).reset_index()

    table_section5_refu = pd.pivot_table(df_refu[df_refu['Вид страхования'] == 'Обязательное страхование гражданской ответственности владельцев транспортных средств'],
                                   index=['ОКАТО', 'Тип страхователя'],
                                   columns=[],
                                   values=["Страховой случай"],
                                   aggfunc=pd.Series.nunique,
                                   fill_value=0).reset_index()


    #Выплаты
    payment = self.folder + '/Убытки.xlsx'
    df_pay = pd.read_excel(payment)
    df_pay.drop([0, 1, 2], inplace=True)

    columns = df_pay.columns.tolist()
    required_col = ['Вид страхования', 'ОКАТО региона', 'ОКАТО территории использо- вания ТС Первоначальный',
                    'Вред жизни', 	'Вред здоровью', 'Тип потерпевшего', 'Риск', 'Риск Причина выплаты',
                    'По реше- нию суда', 'Евро- протокол', 'Страховой случай', 'Заявление', 'Суммы выплаты']

    if (set(required_col) != set(required_col) & set(columns)):
        differences = list(set(required_col) - set(columns))
        append_or_crush = ", ".join(str(element) for element in differences)
        self.error = f'В таблице убытки не хватает столбцов:\n{append_or_crush}'
        raise Exception

    df_pay.insert(loc=len(df_refu.columns), column='Раздел', value=None)
    df_pay.insert(loc=len(df_refu.columns), column='Вид в 162', value=None)
    df_pay = df_pay.set_index("Вид страхования")
    df_pay.update(df_ins_type)
    df_pay = df_pay.reset_index()
    df_pay.insert(3, 'ОКАТО', np.where(df_pay[
                                              'Вид страхования'] != 'Обязательное страхование гражданской ответственности владельцев транспортных средств',
                                          df_pay['ОКАТО региона'] / 1000,
                                          df_pay['ОКАТО территории использо- вания ТС Первоначальный'] / 1000000000))
    df_pay['ОКАТО'] = df_pay['ОКАТО'].astype('int')

    df_pay.insert(9, 'Вид договора', np.where(df_pay[
                                                  'Вид страхования'] == 'Обязательное страхование гражданской ответственности владельцев транспортных средств',
                                              df_pay.apply(type_of_ins, axis=1), None))
    def risk_for_sec1(row):
        if (row['Вид страхования'] == 'Обязательное страхование гражданской ответственности владельцев транспортных средств'):
            if(row['Вред жизни']=='Да'):
                value = 'Причинение вреда жизни третьих лиц'
            elif (row['Вред здоровью'] == 'Да'):
                value = 'Причинение вреда здоровью третьих лиц'
            elif(row['Тип потерпевшего'] == 'Физическое лицо'):
                value = 'Причинение вреда имуществу физических лиц'
            else:
                value = 'Причинение вреда имуществу юридических лиц'
        else:
            if(row['Риск'] == 'залив (аварии отопительных систем, канализационных и водопроводных сетей)'):
                value = 'Аварии отопительных систем, канализационных и водопроводных сетей'
            elif (row['Риск'] == 'ущерб в результате ДТП'):
                value = 'Аварии транспортных средств'
            elif (row['Раздел'] == '03.2. ДМС'):
                value = 'Заболевание'
            elif (row['Риск'] == 'ущерб в результате противоправных действий третьих лиц'):
                value = 'Иные противоправные действия третих лиц'
            elif (row['Риск Причина выплаты'] == 'пожар'):
                value = 'Пожар'
            elif (row['Риск'] == 'причинение вреда имуществу физических лиц'):
                value = 'Причинение вреда имуществу физических лиц'
            elif (row['Риск Причина выплаты'] == 'стихийное бедствие'):
                value = 'Стихийные бедствия'
            elif (row['Риск Причина выплаты'] == 'смерть'):
                value = 'Смерть'
            elif (row['Риск Причина выплаты'] == 'прочее'):
                value = 'Прочие'
            else:
                value = None
        return value
    df_pay.insert(4, 'Риски Раздел1', df_pay.apply(risk_for_sec1, axis=1))

    def risk(row):
        if(row['Риск Причина выплаты'] == 'нанесение ущерба имуществу'and re.fullmatch(r'.*направление.*', row['Риск'])):
            value = 'нанесение ущерба имуществу (ремонт)'
        elif (row['Вид страхования'] == 'Обязательное страхование гражданской ответственности владельцев транспортных средств'):
            value = row['Риск Причина выплаты']
        else:
            value = row['Риск']
        return value
    df_pay.insert(4, 'Риски', df_pay.apply(risk, axis=1))

    df_pay_court = df_pay[df_pay['По реше- нию суда'] == 'Да']
    df_pay_euro = df_pay[df_pay['Евро- протокол'] == 'Да']

    table_section1_pay = pd.pivot_table(df_pay,
                                   index=["Раздел"],
                                   columns=['Тип страхователя'],
                                   values=["Страховой случай", 'Суммы выплаты', 'Заявление'],
                                   aggfunc={"Страховой случай": pd.Series.nunique,'Заявление': pd.Series.nunique, "Суммы выплаты": pd.Series.sum},
                                   fill_value=0).reset_index()

    table_section1_pay_type = pd.pivot_table(df_pay,
                                   index=["Вид в 162"],
                                   columns=['Тип страхователя'],
                                   values=["Страховой случай", 'Суммы выплаты', 'Заявление'],
                                   aggfunc={"Страховой случай": pd.Series.nunique, 'Заявление': pd.Series.nunique, "Суммы выплаты": pd.Series.sum},
                                   fill_value=0).reset_index()

    table_section_2_pay = pd.pivot_table(df_pay[df_pay['Вид страхования'] != 'Обязательное страхование гражданской ответственности владельцев транспортных средств'],
                                           index=["ОКАТО", "Раздел"],
                                           columns=['Тип страхователя'],
                                           values=["Страховой случай", 'Суммы выплаты'],
                                           aggfunc={"Страховой случай": pd.Series.nunique, "Суммы выплаты": pd.Series.sum},
                                           fill_value=0,
                                           margins=True).reset_index()

    table_section5_pay_type = pd.pivot_table(df_pay[df_pay['Вид страхования'] == 'Обязательное страхование гражданской ответственности владельцев транспортных средств'],
                                   index=['ОКАТО', 'Тип страхователя'],
                                   columns=['Вид договора'],
                                   values=["Страховой случай", 'Заявление', 'Суммы выплаты'],
                                   aggfunc={"Страховой случай": pd.Series.nunique, 'Заявление': pd.Series.nunique, "Суммы выплаты": pd.Series.sum},
                                   fill_value=0,
                                   margins=True).reset_index()

    table_section5_pay_risk = pd.pivot_table(df_pay[df_pay['Вид страхования'] == 'Обязательное страхование гражданской ответственности владельцев транспортных средств'],
                                   index=['ОКАТО', 'Тип страхователя'],
                                   columns=['Риски', 'Тип потерпевшего'],
                                   values=['Заявление', 'Суммы выплаты'],
                                   aggfunc={'Заявление': pd.Series.nunique, "Суммы выплаты": pd.Series.sum},
                                   fill_value=0).reset_index()

    table_section1_pay_court = pd.pivot_table(df_pay_court,
                                   index=["Раздел"],
                                   columns=['Тип страхователя'],
                                   values=['Суммы выплаты'],
                                   aggfunc={"Суммы выплаты": pd.Series.sum},
                                   fill_value=0).reset_index()

    table_section1_pay_type_court = pd.pivot_table(df_pay_court,
                                   index=["Вид в 162"],
                                   columns=['Тип страхователя'],
                                   values=['Суммы выплаты'],
                                   aggfunc={"Суммы выплаты": pd.Series.sum},
                                   fill_value=0).reset_index()

    table_section5_pay_court = pd.pivot_table(df_pay_court[df_pay_court['Вид страхования'] == 'Обязательное страхование гражданской ответственности владельцев транспортных средств'],
                                   index=['ОКАТО', 'Тип страхователя'],
                                   columns=[],
                                   values=['Суммы выплаты'],
                                   aggfunc=pd.Series.sum,
                                   fill_value=0).reset_index()

    table_section5_pay_euro = pd.pivot_table(df_pay_euro[df_pay_euro['Вид страхования'] == 'Обязательное страхование гражданской ответственности владельцев транспортных средств'],
                                   index=['ОКАТО', 'Тип страхователя'],
                                   columns=[],
                                   values=['Суммы выплаты', "Страховой случай"],
                                   aggfunc={"Страховой случай": pd.Series.nunique, "Суммы выплаты": pd.Series.sum},
                                   fill_value=0).reset_index()

    table_section1_risk = pd.pivot_table(df_pay,
                                   index=["Раздел", 'Тип страхователя'],
                                   columns=['Риски Раздел1'],
                                   values=['Суммы выплаты', 'Заявление'],
                                   aggfunc={'Заявление': pd.Series.nunique, "Суммы выплаты": pd.Series.sum},
                                   fill_value=0).reset_index()

    table_section1_risk_type = pd.pivot_table(df_pay,
                                   index=["Вид в 162", 'Тип страхователя'],
                                   columns=['Риски Раздел1'],
                                   values=['Суммы выплаты', 'Заявление'],
                                   aggfunc={'Заявление': pd.Series.nunique, "Суммы выплаты": pd.Series.sum},
                                   fill_value=0).reset_index()


    #КВ
    commision = self.folder + '/Комиссионное вознаграждение.xlsx'
    df_com = pd.read_excel(commision)
    df_com.drop(0, inplace=True)

    columns = df_com.columns.tolist()
    required_col = ['Вид страхования', 'Вид агента', 'Размер начисленной комиссии', 'Полис']

    if (set(required_col) != set(required_col) & set(columns)):
        differences = list(set(required_col) - set(columns))
        append_or_crush = ", ".join(str(element) for element in differences)
        self.error = f'В таблице Комиссионное вознаграждение не хватает столбцов:\n{append_or_crush}'
        raise Exception

    df_com.insert(loc=len(df_com.columns), column='Раздел', value=None)
    df_com = df_com.set_index("Вид страхования")
    df_com.update(df_ins_type)
    df_com = df_com.reset_index()


    table_section8_com = pd.pivot_table(df_com,
                                   index=["Раздел"],
                                   columns=['Вид агента'],
                                   values=["Размер начисленной комиссии", 'Полис'],
                                   aggfunc={"Размер начисленной комиссии": pd.Series.sum, 'Полис': pd.Series.nunique},
                                   fill_value=0).reset_index()


    #Перестрахование
    reinsurance_pre = self.folder + '/Перестрахование премии.xlsx'
    df_repre = pd.read_excel(reinsurance_pre)
    df_repre.drop([0, 1 , 2, 3], inplace=True)

    columns = df_repre.columns.tolist()
    required_col = ['Вид страхования', 'Метод передачи риска', 'Суммы начисления премии перестрах.']

    if (set(required_col) != set(required_col) & set(columns)):
        differences = list(set(required_col) - set(columns))
        append_or_crush = ", ".join(str(element) for element in differences)
        self.error = f'В таблице Действующие не хватает столбцов:\n{append_or_crush}'
        raise Exception

    df_repre.insert(loc=len(df_repre.columns), column='Раздел', value=None)
    df_repre = df_repre.set_index("Вид страхования")
    df_repre.update(df_ins_type)
    df_repre = df_repre.reset_index()


    table_section8_repre = pd.pivot_table(df_repre,
                                   index=["Раздел"],
                                   columns=['Метод передачи риска'],
                                   values=["Суммы начисления премии перестрах."],
                                   aggfunc={"Суммы начисления премии перестрах.": pd.Series.sum},
                                   fill_value=0).reset_index()

    reinsurance_pay = self.folder + '/Перестрахование убытки.xlsx'
    df_repay = pd.read_excel(reinsurance_pay)
    df_repay.drop([0, 1, 2, 3], inplace=True)

    df_repay.insert(loc=len(df_repay.columns), column='Раздел', value=None)
    df_repay = df_repay.set_index("Вид страхования")
    df_repay.update(df_ins_type)
    df_repay = df_repay.reset_index()

    columns = df_repay.columns.tolist()
    required_col = ['Вид страхования', 'Метод передачи риска', 'Размер доли перестраховщика убытков']

    if (set(required_col) != set(required_col) & set(columns)):
        differences = list(set(required_col) - set(columns))
        append_or_crush = ", ".join(str(element) for element in differences)
        self.error = f'В таблице Действующие не хватает столбцов:\n{append_or_crush}'
        raise Exception

    table_section8_repay = pd.pivot_table(df_repay,
                                   index=["Раздел"],
                                   columns=['Метод передачи риска'],
                                   values=["Размер доли перестраховщика убытков"],
                                   aggfunc={"Размер доли перестраховщика убытков": pd.Series.sum},
                                   fill_value=0).reset_index()

    #Неустойки
    penalties = self.folder + '/Неустойки.xlsx'
    df_pen = pd.read_excel(penalties)
    df_pen.drop([0, 1, 2], inplace=True)

    columns = df_pen.columns.tolist()
    required_col = ['Вид страхования', 'Статья расходов', 'ОКАТО территории использо- вания ТС Первоначальный',
                    'Заявление', 'Суммы выплаты']

    if (set(required_col) != set(required_col) & set(columns)):
        differences = list(set(required_col) - set(columns))
        append_or_crush = ", ".join(str(element) for element in differences)
        self.error = f'В таблице неустойки не хватает столбцов:\n{append_or_crush}'
        raise Exception

    def penalty(row):
        if(re.fullmatch(r'.*неустойка.*', row['Статья расходов'])):
            value = 'да'
        else:
            value = 'нет'
        return value
    df_pen.insert(4, 'неустойка', df_pen.apply(penalty, axis=1))
    df_pen = df_pen[df_pen['неустойка'] == 'да']

    df_pen.insert(loc=len(df_pen.columns), column='Раздел', value=None)
    df_pen.insert(loc=len(df_pen.columns), column='Вид в 162', value=None)
    df_pen = df_pen.set_index("Вид страхования")
    df_pen.update(df_ins_type)
    df_pen = df_pen.reset_index()
    df_pen.insert(3, 'ОКАТО', np.where(df_pen['Вид страхования'] != 'Обязательное страхование гражданской ответственности владельцев транспортных средств',
                                          df_pen['ОКАТО региона'] / 1000,
                                          df_pen['ОКАТО территории использо- вания ТС Первоначальный'] / 1000000000))
    df_pen['ОКАТО'] = df_pen['ОКАТО'].astype('int')

    table_section1_pen = pd.pivot_table(df_pen,
                                   index=["Раздел"],
                                   columns=['Тип страхователя'],
                                   values=['Заявление', 'Суммы выплаты'],
                                   aggfunc={'Заявление': pd.Series.nunique, "Суммы выплаты": pd.Series.sum},
                                   fill_value=0).reset_index()

    table_section1_pen_type = pd.pivot_table(df_pen,
                                   index=["Вид в 162"],
                                   columns=['Тип страхователя'],
                                   values=['Заявление', 'Суммы выплаты'],
                                   aggfunc={'Заявление': pd.Series.nunique, "Суммы выплаты": pd.Series.sum},
                                   fill_value=0).reset_index()

    table_section5_pen = pd.pivot_table(df_pen[df_pen['Вид страхования'] == 'Обязательное страхование гражданской ответственности владельцев транспортных средств'],
                                   index=['ОКАТО', 'Тип страхователя'],
                                   columns=[],
                                   values=['Заявление', 'Суммы выплаты'],
                                   aggfunc={'Заявление': pd.Series.nunique, "Суммы выплаты": pd.Series.sum},
                                   fill_value=0).reset_index()

    #ЗНУ
    outstanding_claim_reserve = self.folder + '/ЗНУ.xlsx'
    df_ocr = pd.read_excel(outstanding_claim_reserve)

    columns = df_ocr.columns.tolist()
    required_col = ['Вид страхования', 'ОКАТО территории преим.использования ТС первоначального полиса',
                    'Страховое дело', 'Сумма заявленных, но неурегулированных убытков на отчетную дату']

    if (set(required_col) != set(required_col) & set(columns)):
        differences = list(set(required_col) - set(columns))
        append_or_crush = ", ".join(str(element) for element in differences)
        self.error = f'В таблице ЗНУ не хватает столбцов:\n{append_or_crush}'
        raise Exception

    df_ocr.insert(loc=len(df_ocr.columns), column='Раздел', value=None)
    df_ocr.insert(loc=len(df_ocr.columns), column='Вид в 162', value=None)
    df_ocr = df_ocr.set_index("Вид страхования")
    df_ocr.update(df_ins_type)
    df_ocr = df_ocr.reset_index()
    df_ocr.insert(3, 'ОКАТО', np.where(df_ocr['Вид страхования'] != 'Обязательное страхование гражданской ответственности владельцев транспортных средств',
                                          0,
                                          df_ocr['ОКАТО территории преим.использования ТС первоначального полиса'] / 1000000000))
    df_ocr['ОКАТО'] = df_ocr['ОКАТО'].astype('int')

    table_section1_ocr = pd.pivot_table(df_ocr,
                                   index=["Раздел"],
                                   columns=['Тип страхователя'],
                                   values=['Страховое дело', 'Сумма заявленных, но неурегулированных убытков на отчетную дату'],
                                   aggfunc={'Страховое дело': pd.Series.nunique, "Сумма заявленных, но неурегулированных убытков на отчетную дату": pd.Series.sum},
                                   fill_value=0).reset_index()

    table_section1_ocr_type = pd.pivot_table(df_ocr,
                                   index=["Вид в 162"],
                                   columns=['Тип страхователя'],
                                   values=['Страховое дело', 'Сумма заявленных, но неурегулированных убытков на отчетную дату'],
                                   aggfunc={'Страховое дело': pd.Series.nunique, "Сумма заявленных, но неурегулированных убытков на отчетную дату": pd.Series.sum},
                                   fill_value=0).reset_index()

    table_section5_ocr = pd.pivot_table(df_ocr[df_ocr['Вид страхования'] == 'Обязательное страхование гражданской ответственности владельцев транспортных средств'],
                                   index=['ОКАТО', 'Тип страхователя'],
                                   columns=[],
                                   values=['Страховое дело', 'Сумма заявленных, но неурегулированных убытков на отчетную дату'],
                                   aggfunc={'Страховое дело': pd.Series.nunique, "Сумма заявленных, но неурегулированных убытков на отчетную дату": pd.Series.sum},
                                   fill_value=0).reset_index()

    #Действующие
    existing_policies = self.folder + '/Действующие.xlsx'
    df_existing = pd.read_excel(existing_policies)

    columns = df_existing.columns.tolist()
    required_col = ['Вид страхования', 'Код ОКАТО региона', 'Количество застрахованных',
                    'Код ОКАТО', 'Тип страхователя']

    if (set(required_col) != set(required_col) & set(columns)):
        differences = list(set(required_col) - set(columns))
        append_or_crush = ", ".join(str(element) for element in differences)
        self.error = f'В таблице Действующие не хватает столбцов:\n{append_or_crush}'
        raise Exception

    df_existing.insert(loc=len(df_existing.columns), column='Раздел', value=None)
    df_existing.insert(loc=len(df_existing.columns), column='Вид в 162', value=None)
    df_existing = df_existing.set_index("Вид страхования")
    df_existing.update(df_ins_type)
    df_existing = df_existing.reset_index()
    df_existing.insert(43, 'ОКАТО', np.where(df_existing[
                                              'Вид страхования'] != 'Обязательное страхование гражданской ответственности владельцев транспортных средств',
                                          df_existing['Код ОКАТО региона'] / 1000,
                                          df_existing['Код ОКАТО'] / 1000000000))
    df_existing['ОКАТО'] = df_concluded['ОКАТО2'].astype('int')

    df_existing.insert(3, 'Количество договоров', df_existing.apply(policy_count, axis=1))

    df_existing.insert(4, 'Страховая сумма2', df_existing.apply(sum_insured, axis=1))

    df_existing.insert(9, 'Вид договора', np.where(df_existing[
                                                  'Вид страхования'] == 'Обязательное страхование гражданской ответственности владельцев транспортных средств',
                                              df_existing.apply(type_of_ins, axis=1), None))

    table_section1_existing = pd.pivot_table(df_existing,
                                   index=["Раздел"],
                                   columns=['Тип страхователя'],
                                   values=["Количество договоров", 'Страховая сумма2', 'Количество застрахованных'],
                                   aggfunc={"Количество договоров": pd.Series.sum, "Страховая сумма2": pd.Series.sum, "Количество застрахованных": pd.Series.sum},
                                   fill_value=0).reset_index()

    table_section1_existing_type = pd.pivot_table(df_existing,
                                   index=["Вид в 162"],
                                   columns=['Тип страхователя'],
                                   values=["Количество договоров", 'Страховая сумма2', 'Количество застрахованных'],
                                   aggfunc={"Количество договоров": pd.Series.sum, "Страховая сумма2": pd.Series.sum, "Количество застрахованных": pd.Series.sum},
                                   fill_value=0).reset_index()


    table_section5_existing = pd.pivot_table(df_existing[df_existing['Вид страхования'] == 'Обязательное страхование гражданской ответственности владельцев транспортных средств'],
                                   index=['ОКАТО', 'Тип страхователя'],
                                   columns=['Вид договора'],
                                   values=["Количество договоров"],
                                   aggfunc={"Количество договоров": pd.Series.sum},
                                   fill_value=0).reset_index()



    ##### Заполнение формы
    report = 'шаблон_162.xlsx'
    wb = op.load_workbook(report)

    # 1 Раздел. Причины выплаты
    section_1_risk = wb['Раздел 1. Выплаты']

    table_section1_risk.columns = table_section1_risk.columns.droplevel(0)
    table_section1_risk_type.columns = table_section1_risk_type.columns.droplevel(0)
    col = table_section1_risk.columns.tolist()
    col.pop(0)
    col.pop(0)



    for index, row in table_section1_risk.iterrows():
        for i in range (6, 259):
            if (row.values[1] == 'Юридическое лицо'):
                i = i + 260
            for n in range(0, int(len(col)/2)):
                for j in range(3, 67, 2):
                    if((section_1_risk.cell(row = i, column = 2).value) == row.values[0]):
                        if(section_1_risk.cell(row = 3, column = j).value == col[n]):
                            section_1_risk.cell(row = i, column = j).value = row.values[n+2]
                            section_1_risk.cell(row=i, column=j+1).value = row.values[n + int(len(col)/2) + 2]
                            break

    for index, row in table_section1_risk_type.iterrows():
        for i in range (6, 259):
            if (row.values[1] == 'Юридическое лицо'):
                i = i + 260
            for n in range(0, int(len(col)/2)):
                for j in range(3, 67, 2):
                    if((section_1_risk.cell(row = i, column = 2).value) == row.values[0]):
                        if(section_1_risk.cell(row = 3, column = j).value == col[n]):
                            section_1_risk.cell(row = i, column = j).value = row.values[n+2]
                            section_1_risk.cell(row=i, column=j+1).value = row.values[n + int(len(col)/2) + 2]
                            break

     #заполнение
    for x in range(0, 261, 260):
        rows = [147, 154]
        for i in range(3, 67):
            section_1_risk.cell(row=146 + x, column=i).value = 0
            for j in rows:
                if (section_1_risk.cell(row=j + x, column=i).value is None):
                    section_1_risk.cell(row=j + x, column=i).value = 0
                section_1_risk.cell(row=146 + x, column=i).value = section_1_risk.cell(row=146 + x,
                                                                                         column=i).value + section_1_risk.cell(row=j + x, column=i).value
        rows = [160, 163, 164, 165, 168, 169, 180, 184]
        for i in range(3, 67):
            section_1_risk.cell(row=159 + x, column=i).value = 0
            for j in rows:
                if (section_1_risk.cell(row=j + x, column=i).value is None):
                    section_1_risk.cell(row=j + x, column=i).value = 0
                section_1_risk.cell(row=159 + x, column=i).value = section_1_risk.cell(row=159 + x,
                                                                                         column=i).value + section_1_risk.cell(row=j + x, column=i).value
        rows = [189, 193, 194, 195, 198, 202, 205, 211]
        for i in range(3, 67):
            section_1_risk.cell(row=188 + x, column=i).value = 0
            for j in rows:
                if (section_1_risk.cell(row=j + x, column=i).value is None):
                    section_1_risk.cell(row=j + x, column=i).value = 0
                section_1_risk.cell(row=188 + x, column=i).value = section_1_risk.cell(row=188 + x,
                                                                                       column=i).value + section_1_risk.cell(
                    row=j + x, column=i).value
        rows = [159, 188, 221, 222]
        for i in range(3, 67):
            section_1_risk.cell(row=158 + x, column=i).value = 0
            for j in rows:
                if (section_1_risk.cell(row=j + x, column=i).value is None):
                    section_1_risk.cell(row=j + x, column=i).value = 0
                section_1_risk.cell(row=158 + x, column=i).value = section_1_risk.cell(row=158 + x,
                                                                                       column=i).value + section_1_risk.cell(
                    row=j + x, column=i).value

    # 1 Раздел. Премии и выплаты
    section_1_pivot = wb['Раздел 1. Прем. и выпл.']

     #премии
    for index, row in table_section1_pre.iterrows():

        for i in range(7, 260):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=3).value = row.values[2]
                break

        for i in range(266, 519):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=3).value = row.values[3]
                break

        for i in range(525, 778):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=3).value = row.values[1]
                break

    for index, row in table_ins_type_pre.iterrows():

        for i in range(7, 260):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=3).value = row.values[2]
                break

        for i in range(266, 519):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=3).value = row.values[3]
                break

        for i in range(525, 778):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=3).value = row.values[1]
                break

     #заключенные
    for index, row in table_section1_cluded.iterrows():

        for i in range(7, 260):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=4).value = row.values[8]
                section_1_pivot.cell(row=i, column=5).value = row.values[2]
                section_1_pivot.cell(row=i, column=6).value = row.values[11]
                section_1_pivot.cell(row=i, column=7).value = row.values[5]
                break

        for i in range(266, 519):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=4).value = row.values[9]
                section_1_pivot.cell(row=i, column=5).value = row.values[3]
                section_1_pivot.cell(row=i, column=6).value = row.values[12]
                section_1_pivot.cell(row=i, column=7).value = row.values[6]
                break

        for i in range(525, 778):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=4).value = row.values[7]
                section_1_pivot.cell(row=i, column=5).value = row.values[1]
                section_1_pivot.cell(row=i, column=6).value = row.values[10]
                section_1_pivot.cell(row=i, column=7).value = row.values[4]
                break

    for index, row in table_section1_cluded_type.iterrows():

        for i in range(7, 260):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=4).value = row.values[8]
                section_1_pivot.cell(row=i, column=5).value = row.values[2]
                section_1_pivot.cell(row=i, column=6).value = row.values[11]
                section_1_pivot.cell(row=i, column=7).value = row.values[5]
                break

        for i in range(266, 519):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=4).value = row.values[9]
                section_1_pivot.cell(row=i, column=5).value = row.values[3]
                section_1_pivot.cell(row=i, column=6).value = row.values[12]
                section_1_pivot.cell(row=i, column=7).value = row.values[6]
                break

        for i in range(525, 778):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=4).value = row.values[7]
                section_1_pivot.cell(row=i, column=5).value = row.values[1]
                section_1_pivot.cell(row=i, column=6).value = row.values[10]
                section_1_pivot.cell(row=i, column=7).value = row.values[4]
                break


     #возвраты
    for index, row in table_section1_term.iterrows():

        for i in range(7, 260):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=8).value = row.values[3]
                section_1_pivot.cell(row=i, column=9).value = row.values[1]
                break

        for i in range(266, 519):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=8).value = row.values[4]
                section_1_pivot.cell(row=i, column=9).value = row.values[2]
                break

    for index, row in table_section1_term_type.iterrows():

        for i in range(7, 260):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=8).value = - row.values[3]
                section_1_pivot.cell(row=i, column=9).value = row.values[1]
                break

        for i in range(266, 519):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=8).value = - row.values[4]
                section_1_pivot.cell(row=i, column=9).value = row.values[2]
                break


     #действующие
    for index, row in table_section1_existing.iterrows():

        for i in range(7, 260):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=10).value = row.values[2]
                section_1_pivot.cell(row=i, column=11).value = row.values[8]
                section_1_pivot.cell(row=i, column=12).value = row.values[5]
                break

        for i in range(266, 519):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=10).value = row.values[3]
                section_1_pivot.cell(row=i, column=11).value = row.values[9]
                section_1_pivot.cell(row=i, column=12).value = row.values[6]
                break

        for i in range(525, 778):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=10).value = row.values[1]
                section_1_pivot.cell(row=i, column=11).value = row.values[7]
                section_1_pivot.cell(row=i, column=12).value = row.values[4]
                break

    for index, row in table_section1_existing_type.iterrows():

        for i in range(7, 260):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=10).value = row.values[2]
                section_1_pivot.cell(row=i, column=11).value = row.values[8]
                section_1_pivot.cell(row=i, column=12).value = row.values[5]
                break

        for i in range(266, 519):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=10).value = row.values[3]
                section_1_pivot.cell(row=i, column=11).value = row.values[9]
                section_1_pivot.cell(row=i, column=12).value = row.values[6]
                break

        for i in range(525, 778):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=10).value = row.values[1]
                section_1_pivot.cell(row=i, column=11).value = row.values[7]
                section_1_pivot.cell(row=i, column=12).value = row.values[4]
                break


     #заявленные
    for index, row in table_section1_rep.iterrows():

        for i in range(7, 260):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=13).value = row.values[1]
                break

        for i in range(266, 519):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=13).value = row.values[2]
                break



    for index, row in table_section1_rep_type.iterrows():

        for i in range(7, 260):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=13).value = row.values[1]
                break

        for i in range(266, 519):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=13).value = row.values[2]
                break

     #отказы
    for index, row in table_section1_refu.iterrows():

        for i in range(7, 260):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=15).value = row.values[1]
                break

        for i in range(266, 519):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=15).value = row.values[2]
                break

    for index, row in table_section1_refu_type.iterrows():

        for i in range(7, 260):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=15).value = row.values[1]
                break

        for i in range(266, 519):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=15).value = row.values[2]
                break

     #ЗНУ
    for index, row in table_section1_ocr.iterrows():

        for i in range(7, 260):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=16).value = row.values[1]
                section_1_pivot.cell(row=i, column=17).value = row.values[3]
                break

        for i in range(266, 519):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=16).value = row.values[2]
                section_1_pivot.cell(row=i, column=17).value = row.values[4]
                break

    for index, row in table_section1_ocr_type.iterrows():

        for i in range(7, 260):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=16).value = row.values[1]
                section_1_pivot.cell(row=i, column=17).value = row.values[3]
                break

        for i in range(266, 519):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=16).value = row.values[2]
                section_1_pivot.cell(row=i, column=17).value = row.values[4]
                break

     #выплаты
    for index, row in table_section1_pay.iterrows():

        for i in range(7, 260):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=14).value = row.values[3]
                section_1_pivot.cell(row=i, column=18).value = row.values[1]
                section_1_pivot.cell(row=i, column=19).value = row.values[5]
                break

        for i in range(266, 519):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=14).value = row.values[4]
                section_1_pivot.cell(row=i, column=18).value = row.values[2]
                section_1_pivot.cell(row=i, column=19).value = row.values[6]
                break

    for index, row in table_section1_pay_type.iterrows():

        for i in range(7, 260):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=14).value = row.values[3]
                section_1_pivot.cell(row=i, column=18).value = row.values[1]
                section_1_pivot.cell(row=i, column=19).value = row.values[5]
                break

        for i in range(266, 519):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=14).value = row.values[4]
                section_1_pivot.cell(row=i, column=18).value = row.values[2]
                section_1_pivot.cell(row=i, column=19).value = row.values[6]
                break

    for index, row in table_section1_pay_court.iterrows():

        for i in range(7, 260):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=20).value = row.values[1]
                break

    for index, row in table_section1_pay_type_court.iterrows():

        for i in range(7, 260):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=20).value = row.values[1]
                break


     #неустойки
    for index, row in table_section1_pen.iterrows():

        for i in range(7, 260):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=29).value = row.values[1]
                section_1_pivot.cell(row=i, column=30).value = row.values[3]
                break

        for i in range(266, 519):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=29).value = row.values[2]
                section_1_pivot.cell(row=i, column=30).value = row.values[4]
                break

    for index, row in table_section1_pen_type.iterrows():

        for i in range(7, 260):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=29).value = row.values[1]
                section_1_pivot.cell(row=i, column=30).value = row.values[3]
                break

        for i in range(266, 519):

            if (section_1_pivot.cell(row=i, column=2).value == row.values[0]):
                section_1_pivot.cell(row=i, column=29).value = row.values[2]
                section_1_pivot.cell(row=i, column=30).value = row.values[4]
                break


    # Заполнение раздела 1
    for x in range(0, 259*2+1, 259):

        for n in range(1, 4):

            for i in range(3, 35):
                if (section_1_pivot.cell(row=148 + x, column=i).value is None):
                    section_1_pivot.cell(row=148 + x, column=i).value = 0
                if (section_1_pivot.cell(row=155 + x, column=i).value is None):
                    section_1_pivot.cell(row=155 + x, column=i).value = 0
                section_1_pivot.cell(row=147 + x, column=i).value = section_1_pivot.cell(row=148 + x,
                                                                                         column=i).value + section_1_pivot.cell(
                    row=155 + x, column=i).value

        rows = [161, 164, 165, 166, 169, 170, 181, 185]
        for i in range(3, 35):
            section_1_pivot.cell(row=160 + x, column=i).value = 0
            for j in rows:
                if (section_1_pivot.cell(row=j + x, column=i).value is None):
                    section_1_pivot.cell(row=j + x, column=i).value = 0
                section_1_pivot.cell(row=160 + x, column=i).value = section_1_pivot.cell(row=160 + x,
                                                                                         column=i).value + section_1_pivot.cell(
                    row=j + x, column=i).value

        rows = [190, 194, 195, 196, 199, 203, 206, 212]
        for i in range(3, 35):
            section_1_pivot.cell(row=189 + x, column=i).value = 0
            for j in rows:
                if (section_1_pivot.cell(row=j + x, column=i).value is None):
                    section_1_pivot.cell(row=j + x, column=i).value = 0
                section_1_pivot.cell(row=189 + x, column=i).value = section_1_pivot.cell(row=189 + x,
                                                                                         column=i).value + section_1_pivot.cell(
                    row=j + x, column=i).value

        rows = [161, 164, 165, 166, 169, 170, 181, 185]
        for i in range(3, 35):
            section_1_pivot.cell(row=160 + x, column=i).value = 0
            for j in rows:
                if (section_1_pivot.cell(row=j + x, column=i).value is None):
                    section_1_pivot.cell(row=j + x, column=i).value = 0
                section_1_pivot.cell(row=160 + x, column=i).value = section_1_pivot.cell(row=160 + x,
                                                                                         column=i).value + section_1_pivot.cell(
                    row=j + x, column=i).value

        rows = [160, 189, 222, 223]
        for i in range(3, 35):
            section_1_pivot.cell(row=159 + x, column=i).value = 0
            for j in rows:
                if (section_1_pivot.cell(row=j + x, column=i).value is None):
                    section_1_pivot.cell(row=j + x, column=i).value = 0
                section_1_pivot.cell(row=159 + x, column=i).value = section_1_pivot.cell(row=159 + x,
                                                                                         column=i).value + section_1_pivot.cell(
                    row=j + x, column=i).value


    # 2 Раздел. Добровольные виды страхования
    section_2 = wb['Раздел 2. ОКАТО']

    x_04_1 = [1, 2, 3, 4, 5, 6, 12, 13]
    x_04_2 = [1, 4, 5, 6, 7, 8, 9, 10]
    x_04 = [1, 15, 26, 27]
    columns_2 = [c for c in range(3, 13)]

        #премии
    for n in range(4, 6512, 64):
        for index, row in table_okato_section_2_pre.iterrows():
            if(int(section_2.cell(row=n, column=2).value) == row.values[0]):

                for i in range(n + 6, n + 44):

                    if (section_2.cell(row=i, column=2).value == row.values[1]):
                        section_2.cell(row=i, column=3).value = row.values[5]
                        section_2.cell(row=i, column=4).value = row.values[3]
                        break

        #заключенные
    for n in range(4, 6512, 64):
        for index, row in table_section_2_cluded.iterrows():
            if (int(section_2.cell(row=n, column=2).value) == row.values[0]):

                for i in range(n + 6, n + 44):

                    if (section_2.cell(row=i, column=2).value == row.values[1]):
                        section_2.cell(row=i, column=5).value = row.values[5]
                        section_2.cell(row=i, column=6).value = row.values[3]
                        section_2.cell(row=i, column=7).value = row.values[9]
                        section_2.cell(row=i, column=8).value = row.values[7]
                        break

        #убытки
    for n in range(4, 6512, 64):
        for index, row in table_section_2_pay.iterrows():
            if (int(section_2.cell(row=n, column=2).value) == row.values[0]):

                for i in range(n + 6, n + 44):

                    if (section_2.cell(row=i, column=2).value == row.values[1]):
                        section_2.cell(row=i, column=9).value = row.values[7]
                        section_2.cell(row=i, column=10).value = row.values[5]
                        section_2.cell(row=i, column=11).value = row.values[4]
                        section_2.cell(row=i, column=12).value = row.values[2]
                        break


    #заполнение раздела 2
    for n in range(4, 6512, 64):
        for k in columns_2:
            if (section_2.cell(row=n + 9, column=k).value is None):
                section_2.cell(row=n + 9, column=k).value = 0
            if (section_2.cell(row=n + 12, column=k).value is None):
                section_2.cell(row=n + 12, column=k).value = 0
            section_2.cell(row=n + 8, column=k).value = section_2.cell(row=n + 9, column=k).value + section_2.cell(
                row=n + 12, column=k).value

            if (section_2.cell(row=n + 14, column=k).value is None):
                section_2.cell(row=n + 14, column=k).value = 0
            for j in x_04_1:
                if (section_2.cell(row=n + j + 14, column=k).value is None):
                    section_2.cell(row=n + j + 14, column=k).value = 0
                section_2.cell(row=n + 14, column=k).value = section_2.cell(row=n + 14,
                                                                            column=k).value + section_2.cell(
                    row=n + j + 14, column=k).value

            if (section_2.cell(row=n + 28, column=k).value is None):
                section_2.cell(row=n + 28, column=k).value = 0
            for j in x_04_2:

                if (section_2.cell(row=n + j + 28, column=k).value is None):
                    section_2.cell(row=n + j + 28, column=k).value = 0
                section_2.cell(row=n + 28, column=k).value = section_2.cell(row=n + 28,
                                                                            column=k).value + section_2.cell(
                    row=n + j + 28, column=k).value

            if (section_2.cell(row=n + 13, column=k).value is None):
                section_2.cell(row=n + 13, column=k).value = 0
            for j in x_04:
                if (section_2.cell(row=n + j + 13, column=k).value is None):
                    section_2.cell(row=n + j + 13, column=k).value = 0
                section_2.cell(row=n + 13, column=k).value = section_2.cell(row=n + 13,
                                                                            column=k).value + section_2.cell(
                    row=n + j + 13, column=k).value


    # 5 Раздел. ОСАГО
    section_5 = wb['Раздел 5. ОСАГО в разрезе ОКАТО']

        #премии
    for index, row in table_section5_pre.iterrows():
        for i in range(9, 186, 2):

            if (int(section_5.cell(row=i, column=2).value) == row.values[0]):
                if (row.values[1] == 'Юридическое лицо'):
                    i = i+1
                section_5.cell(row=i, column=5).value = row.values[2]
                section_5.cell(row=i, column=6).value = row.values[3]
                break


        #заключенные
    for index, row in table_section5_cluded.iterrows():
        for i in range(9, 186, 2):

            if (int(section_5.cell(row=i, column=2).value) == row.values[0]):
                if (row.values[1] == 'Юридическое лицо'):
                    i = i+1
                section_5.cell(row=i, column=9).value = row.values[4]
                section_5.cell(row=i, column=10).value = row.values[5]
                section_5.cell(row=i, column=13).value = row.values[2]
                section_5.cell(row=i, column=14).value = row.values[3]
                break

        #возвраты
        #возвраты
    for index, row in table_section5_term.iterrows():
        for i in range(9, 186, 2):

            if (int(section_5.cell(row=i, column=2).value) == row.values[0]):
                if (row.values[1] == 'Юридическое лицо'):
                    i = i+1
                section_5.cell(row=i, column=18).value = - row.values[3]
                section_5.cell(row=i, column=19).value = row.values[2]
                break

    #Действующие
    for index, row in table_section5_existing.iterrows():
        for i in range(9, 186, 2):

            if (int(section_5.cell(row=i, column=2).value) == row.values[0]):
                if (row.values[1] == 'Юридическое лицо'):
                    i = i+1
                section_5.cell(row=i, column=20).value = row.values[2]
                section_5.cell(row=i, column=21).value = row.values[3]
                break

    #заявленные
    for index, row in table_section5_rep.iterrows():
        for i in range(9, 186, 2):

            if (int(section_5.cell(row=i, column=2).value) == row.values[0]):
                if (row.values[1] == 'Юридическое лицо'):
                    i = i+1
                section_5.cell(row=i, column=24).value = row.values[2]
                break

    #ЗНУ
    for index, row in table_section5_ocr.iterrows():
        for i in range(9, 186, 2):

            if (int(section_5.cell(row=i, column=2).value) == row.values[0]):
                if (row.values[1] == 'Юридическое лицо'):
                    i = i+1
                section_5.cell(row=i, column=27).value = row.values[2]
                section_5.cell(row=i, column=28).value = row.values[3]
                break

    #убытки
    for index, row in table_section5_pay_type.iterrows():
        for i in range(9, 186, 2):

            if (int(section_5.cell(row=i, column=2).value) == row.values[0]):
                if (row.values[1] == 'Юридическое лицо'):
                    i = i+1
                section_5.cell(row=i, column=25).value = row.values[9]
                section_5.cell(row=i, column=29).value = row.values[2]
                section_5.cell(row=i, column=30).value = row.values[3] + row.values[4]
                section_5.cell(row=i, column=31).value = row.values[3]
                section_5.cell(row=i, column=33).value = row.values[10]
                section_5.cell(row=i, column=34).value = row.values[11] + row.values[12]
                section_5.cell(row=i, column=35).value = row.values[11]
                break

    for index, row in table_section5_pay_risk.iterrows():
        for i in range(9, 186, 2):

            if (int(section_5.cell(row=i, column=2).value) == row.values[0]):
                if (row.values[1] == 'Юридическое лицо'):
                    i = i+1
                section_5.cell(row=i, column=37).value = row.values[2]
                section_5.cell(row=i, column=38).value = row.values[7]
                section_5.cell(row=i, column=39).value = row.values[3]
                section_5.cell(row=i, column=40).value = row.values[8]
                section_5.cell(row=i, column=41).value = row.values[4] + row.values[6]
                section_5.cell(row=i, column=42).value = row.values[10] + row.values[12]
                section_5.cell(row=i, column=43).value = row.values[5] + row.values[7]
                section_5.cell(row=i, column=44).value = row.values[11] + row.values[13]
                section_5.cell(row=i, column=45).value = row.values[6]
                section_5.cell(row=i, column=46).value = row.values[12]
                section_5.cell(row=i, column=47).value = row.values[7]
                section_5.cell(row=i, column=48).value = row.values[13]
                break

    for index, row in table_section5_pay_court.iterrows():
        for i in range(9, 186, 2):

            if (int(section_5.cell(row=i, column=2).value) == row.values[0]):
                if (row.values[1] == 'Юридическое лицо'):
                    i = i+1
                section_5.cell(row=i, column=49).value = row.values[2]
                break

    for index, row in table_section5_pay_euro.iterrows():
        for i in range(9, 186, 2):

            if (int(section_5.cell(row=i, column=2).value) == row.values[0]):
                if (row.values[1] == 'Юридическое лицо'):
                    i = i+1
                section_5.cell(row=i, column=50).value = row.values[2]
                section_5.cell(row=i, column=51).value = row.values[3]
                break

    #Неустойки
    for index, row in table_section5_pen.iterrows():
        for i in range(9, 186, 2):

            if (int(section_5.cell(row=i, column=2).value) == row.values[0]):
                if (row.values[1] == 'Юридическое лицо'):
                    i = i+1
                section_5.cell(row=i, column=56).value = row.values[2]
                section_5.cell(row=i, column=57).value = row.values[3]
                break


    table_section8_com.to_excel('Проверка.xlsx')
    #7 Раздел
    section_7 = wb['Раздел 7']

    for index, row in table_section8_com.iterrows():
        for i in range(7, 68):

            if (section_7.cell(row=i, column=2).value == row.values[0]):
                section_7.cell(row=i, column=3).value = row.values[3] + row.values[4]
                section_7.cell(row=i, column=4).value = row.values[3]
                section_7.cell(row=i, column=5).value = row.values[5]
                section_7.cell(row=i, column=6).value = row.values[1] + row.values[2]
                section_7.cell(row=i, column=7).value = row.values[8] + row.values[9]
                section_7.cell(row=i, column=8).value = row.values[8]
                section_7.cell(row=i, column=9).value = row.values[10]
                section_7.cell(row=i, column=10).value = row.values[6] + row.values[7]
                break

    rows = [38, 41]
    for i in range(3, 10):
            section_7.cell(row=37, column=i).value = 0
            for j in rows:
                if (section_7.cell(row=j, column=i).value is None):
                    section_7.cell(row=j, column=i).value = 0
                section_7.cell(row=37, column=i).value = section_7.cell(row=37, column=i).value + section_7.cell(row=j, column=i).value

    rows = [44, 45, 46, 47, 48, 49, 51, 52]
    for i in range(3, 10):
            section_7.cell(row=43, column=i).value = 0
            for j in rows:
                if (section_7.cell(row=j, column=i).value is None):
                    section_7.cell(row=j, column=i).value = 0
                section_7.cell(row=43, column=i).value = section_7.cell(row=43, column=i).value + section_7.cell(row=j, column=i).value

    rows = [54, 57, 58, 59, 60, 61, 62, 63]
    for i in range(3, 10):
            section_7.cell(row=53, column=i).value = 0
            for j in rows:
                if (section_7.cell(row=j, column=i).value is None):
                    section_7.cell(row=j, column=i).value = 0
                section_7.cell(row=53, column=i).value = section_7.cell(row=53, column=i).value + section_7.cell(row=j, column=i).value

    rows = [43, 53, 64, 65]
    for i in range(3, 10):
            section_7.cell(row=42, column=i).value = 0
            for j in rows:
                if (section_7.cell(row=j, column=i).value is None):
                    section_7.cell(row=j, column=i).value = 0
                section_7.cell(row=42, column=i).value = section_7.cell(row=42, column=i).value + section_7.cell(row=j, column=i).value




    #8 Раздел
    section_8 = wb['Раздел 8']

    for index, row in table_section8_repre.iterrows():
        for i in range(8, 71):

            if (section_8.cell(row=i, column=2).value == row.values[0]):
                section_8.cell(row=i, column=3).value = row.values[2]
                section_8.cell(row=i, column=5).value = row.values[1]
                section_8.cell(row=i, column=7).value = row.values[3]
                break

    for index, row in table_section8_repay.iterrows():
        for i in range(8, 71):

            if (section_8.cell(row=i, column=2).value == row.values[0]):
                section_8.cell(row=i, column=24).value = row.values[1]
                break


    rows = [39, 42]
    for i in range(3, 36):
            section_8.cell(row=38, column=i).value = 0
            for j in rows:
                if (section_8.cell(row=j, column=i).value is None):
                    section_8.cell(row=j, column=i).value = 0
                section_8.cell(row=38, column=i).value = section_8.cell(row=38, column=i).value + section_8.cell(row=j, column=i).value

    rows = [45, 46, 47, 48, 49, 50, 52, 53]
    for i in range(3, 36):
            section_8.cell(row=44, column=i).value = 0
            for j in rows:
                if (section_8.cell(row=j, column=i).value is None):
                    section_8.cell(row=j, column=i).value = 0
                section_8.cell(row=44, column=i).value = section_8.cell(row=44, column=i).value + section_8.cell(row=j, column=i).value

    rows = [55, 58, 59, 60, 61, 62, 63, 64]
    for i in range(3, 36):
            section_8.cell(row=54, column=i).value = 0
            for j in rows:
                if (section_8.cell(row=j, column=i).value is None):
                    section_8.cell(row=j, column=i).value = 0
                section_8.cell(row=54, column=i).value = section_8.cell(row=54, column=i).value + section_8.cell(row=j, column=i).value

    rows = [44, 54, 65, 66]
    for i in range(3, 36):
            section_8.cell(row=43, column=i).value = 0
            for j in rows:
                if (section_8.cell(row=j, column=i).value is None):
                    section_8.cell(row=j, column=i).value = 0
                section_8.cell(row=43, column=i).value = section_8.cell(row=43, column=i).value + section_8.cell(row=j, column=i).value



    wb.save('!Форма 162.xlsx')
    import subprocess
    subprocess.call('!Форма 162.xlsx', shell=True)
