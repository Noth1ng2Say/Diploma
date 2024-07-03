import openpyxl as op
import re #регулярные
import os


class Sales_report:
    error = ''
    def __init__(self, folder):
        self.folder = folder


    def form(self):
        directory = self.folder
        files = os.listdir(directory)
        required_files = ['еженедельный отчет.xlsx', 'Казань.xlsx', 'Краснодар.xlsx', 'Москва.xlsx', 'Набережные Челны.xlsx', 'Нижний Новгород.xlsx', 'Самара.xlsx']


        if (sorted(list(set(required_files))) != sorted(list(set(required_files) & set(files)))):
            differences = list(set(required_files) - set(files))
            append_or_crush = ", ".join(str(element) for element in differences)
            self.error = f'В директории не хватает файлов: {append_or_crush}'
            raise Exception



        filename = self.folder + '/еженедельный отчет.xlsx'
        wb = op.load_workbook(filename)
        sheet = wb.active
        max_row = sheet.max_row


        salesReport = 'Шаблон_Отчет по продажам.xlsx'
        nwb = op.load_workbook(salesReport)
        sheetSR = nwb.active
        max_rows = sheetSR.max_row

        #Список филиалов
        branches = []
        for i in range (5, max_rows):
         if (re.fullmatch(r'\D+.*', sheetSR.cell(row = i, column = 1).value)):
                branches.append(sheetSR.cell(row = i, column = 1).value)
        branches.append('Е-Гарант')
        branches.append('Е-ОСАГО сайт')

        #Для Е-осаго
        sheetSR['C90'] = 0
        sheetSR['E90'] = 0
        sheetSR['G90'] = 0
        sheetSR['H90'] = 0

        #Для Закрытых
        for i in range (92, 105):
            sheetSR.cell(row=i, column=5).value = 0

        for i in range (3, max_row + 1):
            branch = sheet.cell(row=i, column=1).value
            if(branch == 'Е-Гарант' or branch == 'Е-ОСАГО сайт'):

                if (sheet.cell(row = i, column = 9).value) is None:
                    sheet.cell(row=i, column=9).value = 0
                if (sheet.cell(row = i, column = 3).value) is None:
                    sheet.cell(row=i, column=3).value = 0
                if (sheet.cell(row = i, column = 8).value) is None:
                    sheet.cell(row=i, column=8).value = 0
                if (sheet.cell(row = i, column = 10).value) is None:
                    sheet.cell(row=i, column=10).value = 0

                sheetSR['C90'] = (sheetSR['C90'].value + sheet.cell(row = i, column = 9).value)
                sheetSR['E90'] = (sheetSR['E90'].value + sheet.cell(row=i, column=3).value)
                sheetSR['G90'] = sheetSR['G90'].value + sheet.cell(row=i, column=8).value
                sheetSR['H90'] = sheetSR['H90'].value + sheet.cell(row=i, column=10).value

            for n in range (5, max_rows - 16):

             if (re.fullmatch(r'\D+.*', branch) and sheetSR.cell(row = n, column = 1).value == branch):

                j = 1

                while (re.fullmatch(r'\d.*', sheet.cell(row = i + j, column = 1).value)):
                    for z in range(15):

                     if (sheetSR.cell(row = n + z, column = 1).value == sheet.cell(row = i + j, column = 1).value):

                        payment = sheet.cell(row = i + j, column = 3).value
                        if payment is None:
                            payment=0
                        else:
                            sheetSR[f'E{n + z}'] = payment / 1000

                        numOfConcludedC0ntracts = sheet.cell(row=i + j, column=8).value
                        if numOfConcludedC0ntracts is None:
                            numOfConcludedC0ntracts = 0
                        else:
                            sheetSR[f'G{n+z}'] = numOfConcludedC0ntracts

                        avgPremium = sheet.cell(row=i + j, column=10).value
                        if avgPremium is None:
                            avgPremium = 0
                        else:
                            sheetSR[f'H{n+z}'] = avgPremium
                    j = j + 1
        #Е-ОСАГО
        sheetSR['C90'] = sheetSR['C90'].value / 1000
        sheetSR['E90'] = sheetSR['E90'].value / 1000

        #Закрытые
        for i in range(3, max_row):

            if (re.fullmatch(r'\D+.*', sheet.cell(row=i, column=1).value) and sheet.cell(row=i, column=1).value not in branches):
                j = 1

                while (re.fullmatch(r'\d.*', sheet.cell(row=i + j, column=1).value)):
                    for z in range(15):

                        if (sheetSR.cell(row=91 + z, column=1).value == sheet.cell(row=i + j, column=1).value):

                            if sheet.cell(row=i + j, column=3).value is None:
                                continue
                            sheetSR.cell(row=91 + z, column=5).value = sheetSR.cell(row=91 + z, column=5).value + sheet.cell(row=i + j, column=3).value / 1000
                    j = j + 1

        # список файлов в директории
        files.remove('еженедельный отчет.xlsx')

        #премии
        for i in range (5, max_rows - 16):
            for file in files:
                if(sheetSR.cell(row = i, column = 1).value == file[:-5]):
                    load_file = self.folder + f'/{file}'
                    lwb = op.load_workbook(load_file)
                    active_sheet = lwb.active
                    for z in range(13):
                        sheetSR.cell(row=i+z+1, column=3).value = active_sheet.cell(row=z+9, column=3).value

        nwb.save('!Отчет по продажам.xlsx')
        import subprocess
        subprocess.call('!Отчет по продажам.xlsx', shell=True)

